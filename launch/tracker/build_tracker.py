#!/usr/bin/env python3
"""
Pick1 — Master Tracker generator.

Produces `MASTER_TRACKER.xlsx` next to this script. Nine tabs covering
everything we plan and measure during the 18-day pre-launch sprint:

  1. KPI Dashboard       — top-level numbers, day-by-day
  2. Content Calendar    — what's getting posted where + perf
  3. Ad Performance      — Meta + TikTok Spark spend / CPL / conversions
  4. Channel ROI         — which channel is winning, % of total
                           (42 channels tracked)
  5. Activation Pipeline — the 22 viral activations + their state
                           (web3 quests, Founder Pass, etc.)
  6. Influencer Outreach — 20-DM micro-influencer trade program
  7. Pick Performance    — daily pick record (the receipts wall data)
  8. $100K Scenario      — what the marketing spend would look like
                           with 100x the budget (street TikToks,
                           macro influencers, PR firm, etc.)
  9. Daily Strategy      — ⭐ THE operational tab. One row per day
                           from May 14 → May 31. Auto-cron vs.
                           user-posts vs. outreach vs. activations
                           launching, with daily + cumulative
                           signup targets. Start here every morning.
 10. Content Engine      — Daily content production blueprint:
                           24-hour post schedule, repurposing matrix
                           (1 piece → 6-8 outlets), hashtag library,
                           time-of-day map per platform.
 11. Aggressive Tactics  — 50 free moves ranked by ROI. CMO playbook
                           with tier-coded tactics (T1 = do all; T5
                           = guerrilla picks). Read alongside
                           launch/CMO_PLAYBOOK.md for the full memo.
 12. $25K Active Plan    — ⭐ CURRENT BUDGET. Full $25K allocation,
                           week-by-week paid cadence, immediate
                           first moves, expected outcomes vs $1K
                           and $100K, risk callouts. Companion
                           doc: launch/AT_25K.md.

Re-run any time to regenerate from scratch:
    python3 launch/tracker/build_tracker.py

Designed so the user can drop it into Google Sheets (File → Import →
Replace spreadsheet) and have a working dashboard immediately. All tabs
have header styling + frozen header rows + conditional formatting where
helpful (CPL traffic-light, win/loss color, etc.).
"""

from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.utils import get_column_letter
from pathlib import Path

# ── Brand styling ────────────────────────────────────────────────────────
BRAND_BG       = "0A0B0D"
BRAND_ACCENT   = "D4FF3A"
BRAND_LINE     = "1A1D22"
BRAND_MUTED    = "9095A0"
WHITE          = "FFFFFF"

HEADER_FONT  = Font(name="Calibri", bold=True, size=11, color=BRAND_BG)
HEADER_FILL  = PatternFill("solid", fgColor=BRAND_ACCENT)
SECTION_FONT = Font(name="Calibri", bold=True, size=14, color=WHITE)
SECTION_FILL = PatternFill("solid", fgColor=BRAND_BG)
BODY_FONT    = Font(name="Calibri", size=10)
SUBTLE_FONT  = Font(name="Calibri", size=9, italic=True, color="666666")
THIN_BORDER  = Border(*[Side(style="thin", color="DDDDDD")] * 4)
CENTER       = Alignment(horizontal="center", vertical="center")
LEFT         = Alignment(horizontal="left",   vertical="center")
WRAP         = Alignment(horizontal="left",   vertical="top", wrap_text=True)


def style_header_row(ws, row, ncols):
    """Apply brand-coloured header styling to a row."""
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 24


def autosize(ws, widths):
    """Apply column widths (list of ints)."""
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Tab 1: KPI Dashboard ─────────────────────────────────────────────────
# One row per day from May 13 (today) through May 31 (launch). The cells
# the founder will fill in nightly are: signups_total, signups_today,
# meta_spend, meta_signups, organic_signups, ig_followers, tiktok_followers,
# email_opens, email_ctr, referral_shares. Everything else (CPL, viral
# coefficient, % of target) is formulas that calculate automatically.
def build_kpi_dashboard(wb):
    ws = wb.create_sheet("1 · KPI Dashboard")
    ws.sheet_view.showGridLines = False

    # Section banner
    ws.merge_cells("A1:M1")
    ws["A1"] = "PICK1 · DAILY KPI DASHBOARD (May 13 → May 31)"
    ws["A1"].font = SECTION_FONT
    ws["A1"].fill = SECTION_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 32

    # Subhead with the north-star target
    ws.merge_cells("A2:M2")
    ws["A2"] = "North Star: 1,500 waitlist signups by May 31  ·  Budget: $1,000"
    ws["A2"].font = SUBTLE_FONT
    ws["A2"].alignment = CENTER

    # Header row 3
    headers = [
        "Day", "Date", "Signups (cumulative)", "Signups (today)",
        "Meta $ today", "Meta signups today", "Meta CPL",
        "Organic signups today", "IG followers", "TikTok followers",
        "Email open %", "Referral shares", "Notes",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h)
    style_header_row(ws, 3, len(headers))
    ws.freeze_panes = "A4"

    # Body rows — one per day
    start = date(2026, 5, 13)
    end   = date(2026, 5, 31)
    days  = (end - start).days + 1
    for i in range(days):
        d = start + timedelta(days=i)
        r = 4 + i
        ws.cell(row=r, column=1, value=f"Day {i+1}")
        ws.cell(row=r, column=2, value=d).number_format = "yyyy-mm-dd"
        # Seed today's known number (Diego = 1 cumulative on Day 1)
        if i == 0:
            ws.cell(row=r, column=3, value=1)
            ws.cell(row=r, column=4, value=1)
        # CPL formula: meta $ / meta signups (when > 0)
        ws.cell(row=r, column=7,
                value=f"=IFERROR(E{r}/F{r},\"\")").number_format = "$#,##0.00"

    # Conditional formatting — CPL traffic light
    cpl_range = f"G4:G{4 + days - 1}"
    ws.conditional_formatting.add(cpl_range,
        CellIsRule(operator="lessThan", formula=["1.5"],
                   fill=PatternFill("solid", fgColor="C6EFCE")))
    ws.conditional_formatting.add(cpl_range,
        CellIsRule(operator="between", formula=["1.5", "3"],
                   fill=PatternFill("solid", fgColor="FFEB9C")))
    ws.conditional_formatting.add(cpl_range,
        CellIsRule(operator="greaterThan", formula=["3"],
                   fill=PatternFill("solid", fgColor="FFC7CE")))

    # Cumulative signups colour scale (progress toward 1,500)
    cum_range = f"C4:C{4 + days - 1}"
    ws.conditional_formatting.add(cum_range,
        ColorScaleRule(start_type="num", start_value=0,    start_color="FFC7CE",
                       mid_type="num",   mid_value=750,    mid_color="FFEB9C",
                       end_type="num",   end_value=1500,   end_color="C6EFCE"))

    autosize(ws, [8, 12, 18, 16, 13, 18, 12, 22, 14, 18, 14, 17, 35])


# ── Tab 2: Content Calendar ──────────────────────────────────────────────
def build_content_calendar(wb):
    ws = wb.create_sheet("2 · Content Calendar")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:K1")
    ws["A1"] = "CONTENT CALENDAR — every post across every channel"
    ws["A1"].font = SECTION_FONT; ws["A1"].fill = SECTION_FILL; ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 32

    headers = [
        "Date", "Channel", "Format", "Pillar", "Topic / Hook",
        "Status", "URL", "Views", "Likes", "Shares", "Signups attributed",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=2, column=c, value=h)
    style_header_row(ws, 2, len(headers))
    ws.freeze_panes = "A3"

    # Seed Day 1 content already planned
    seed_rows = [
        ("2026-05-13", "TikTok",  "Reel",     "Receipt",   "Yesterday 2-1 (SAS 126-97 lock)",       "Drafted",  "", 0, 0, 0, 0),
        ("2026-05-13", "TikTok",  "Reel",     "Daily Pick","Tonight's lock: MCI 81.2% confidence",  "Drafted",  "", 0, 0, 0, 0),
        ("2026-05-13", "IG",      "Static",   "Receipt",   "Post #2: SAS 126-97 bold static",       "To render","", 0, 0, 0, 0),
        ("2026-05-13", "IG",      "Story",    "Daily Pick","Pick locked 🔒 (link sticker)",         "Drafted",  "", 0, 0, 0, 0),
        ("2026-05-13", "Twitter", "Tweet",    "Receipt",   "Yesterday: SAS 126-97 + tonight tease", "Drafted",  "", 0, 0, 0, 0),
        ("2026-05-13", "Twitter", "Tweet",    "Daily Pick","3pm: MCI lock reveal",                  "Scheduled","", 0, 0, 0, 0),
        ("2026-05-13", "Reddit",  "Comment",  "Receipt",   "r/sportsbook daily-talk thread",        "Drafted",  "", 0, 0, 0, 0),
        ("2026-05-13", "Meta",    "Static Ad","Receipt",   "Ad 1: 'Yesterday: Spurs 126-97 ✓'",     "To render","", 0, 0, 0, 0),
        ("2026-05-13", "Meta",    "Static Ad","Daily Pick","Ad 2: 'Tonight's lock 81.2%'",          "To render","", 0, 0, 0, 0),
        ("2026-05-13", "Meta",    "Static Ad","Positioning","Ad 3: 'One pick · 9 sports'",          "To render","", 0, 0, 0, 0),
    ]
    for r, row in enumerate(seed_rows, start=3):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            if c == 1:
                cell.number_format = "yyyy-mm-dd"

    # Status colour coding
    status_range = f"F3:F{3 + len(seed_rows) - 1 + 200}"  # room for growth
    ws.conditional_formatting.add(status_range,
        CellIsRule(operator="equal", formula=['"Posted"'],
                   fill=PatternFill("solid", fgColor="C6EFCE")))
    ws.conditional_formatting.add(status_range,
        CellIsRule(operator="equal", formula=['"Scheduled"'],
                   fill=PatternFill("solid", fgColor="FFEB9C")))
    ws.conditional_formatting.add(status_range,
        CellIsRule(operator="equal", formula=['"Drafted"'],
                   fill=PatternFill("solid", fgColor="DDEBF7")))
    ws.conditional_formatting.add(status_range,
        CellIsRule(operator="equal", formula=['"To render"'],
                   fill=PatternFill("solid", fgColor="FCE4D6")))

    autosize(ws, [12, 10, 11, 13, 42, 11, 28, 9, 9, 9, 18])


# ── Tab 3: Ad Performance ────────────────────────────────────────────────
def build_ad_performance(wb):
    ws = wb.create_sheet("3 · Ad Performance")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    ws["A1"] = "AD PERFORMANCE — Meta + TikTok Spark, day-by-day"
    ws["A1"].font = SECTION_FONT; ws["A1"].fill = SECTION_FILL; ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 32

    headers = [
        "Date", "Platform", "Campaign", "Ad / Creative",
        "Spend", "Impressions", "Clicks", "CPC", "Conversions", "CPL",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=2, column=c, value=h)
    style_header_row(ws, 2, len(headers))
    ws.freeze_panes = "A3"

    # Pre-seed Day 1 with empty Campaign 1 ad rows
    for i, ad in enumerate(["Receipts", "Lock", "Positioning"], start=3):
        ws.cell(row=i, column=1, value=date(2026, 5, 13)).number_format = "yyyy-mm-dd"
        ws.cell(row=i, column=2, value="Meta")
        ws.cell(row=i, column=3, value="LEAD-US-broad-20260513")
        ws.cell(row=i, column=4, value=f"Ad: {ad}")
        # CPC formula
        ws.cell(row=i, column=8, value=f"=IFERROR(E{i}/G{i},\"\")").number_format = "$#,##0.00"
        # CPL formula
        ws.cell(row=i, column=10, value=f"=IFERROR(E{i}/I{i},\"\")").number_format = "$#,##0.00"

    # Conditional formatting on CPL (G column → row 3 onwards)
    ws.conditional_formatting.add("J3:J500",
        CellIsRule(operator="lessThan", formula=["1.5"],
                   fill=PatternFill("solid", fgColor="C6EFCE")))
    ws.conditional_formatting.add("J3:J500",
        CellIsRule(operator="greaterThan", formula=["3"],
                   fill=PatternFill("solid", fgColor="FFC7CE")))

    autosize(ws, [12, 10, 28, 25, 10, 13, 9, 10, 13, 10])


# ── Tab 4: Channel ROI ───────────────────────────────────────────────────
def build_channel_roi(wb):
    ws = wb.create_sheet("4 · Channel ROI")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    ws["A1"] = "CHANNEL ROI — where are signups actually coming from?"
    ws["A1"].font = SECTION_FONT; ws["A1"].fill = SECTION_FILL; ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 32

    headers = ["Channel", "$ Spent", "Signups", "CPL", "% of total signups", "Notes"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=2, column=c, value=h)
    style_header_row(ws, 2, len(headers))
    ws.freeze_panes = "A3"

    channels = ["Meta Ads (cold)", "Meta Ads (retargeting)",
                "TikTok organic", "TikTok Spark (paid boost)",
                "TikTok #BeatPick1 challenge",
                "X / Twitter organic", "Twitter Daily Drop",
                "Crypto Twitter shilling (paid)",
                "Instagram organic", "Instagram bio link",
                "Instagram Story takeovers",
                "Facebook Page organic",
                "Reddit r/SideProject", "Reddit r/giveaways",
                "Reddit r/contests", "Reddit r/sportsbook",
                "Reddit niche-sport subs",
                "Hacker News (Show HN)", "Indie Hackers milestone",
                "LinkedIn (founder post)",
                "Influencer trades (organic)", "Influencer (paid)",
                "YouTube creator sponsors",
                "ProductHunt launch", "Founder essay drop",
                "Email — welcome CTA (Resend)", "Email — daily pick CTA (Resend)",
                "Email — blast to existing list",
                "Gleam sweepstakes (organic)",
                "KingSumo viral giveaway",
                "SparkLoop newsletter cross-promo",
                "Sweep aggregator submissions",
                "AI directory submissions",
                # Web3 channels (added May 14)
                "Zealy quest 'Pick1 Sprint'",
                "Galxe sports quest",
                "Layer3 quest",
                "TaskOn quest (APAC)",
                "Pick1 Discord server",
                "Pick1 Founder Pass (collectible)",
                "Prediction-market community seeding",
                "Daily lineup card shares",
                "Direct / unknown"]
    for i, ch in enumerate(channels, start=3):
        ws.cell(row=i, column=1, value=ch)
        # CPL formula
        ws.cell(row=i, column=4, value=f"=IFERROR(B{i}/C{i},\"\")").number_format = "$#,##0.00"
        # % of total — references the SUM of all signups in column C
        ws.cell(row=i, column=5,
                value=f"=IFERROR(C{i}/SUM($C$3:$C${3 + len(channels) - 1}),\"\")").number_format = "0.0%"

    autosize(ws, [28, 10, 11, 9, 20, 40])


# ── Tab 5: Activation Pipeline ───────────────────────────────────────────
def build_activations(wb):
    ws = wb.create_sheet("5 · Activations")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    ws["A1"] = "VIRAL ACTIVATIONS — the 5 plays + status"
    ws["A1"].font = SECTION_FONT; ws["A1"].fill = SECTION_FILL; ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 32

    headers = ["#", "Activation", "Target launch", "Status", "Owner", "Notes"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=2, column=c, value=h)
    style_header_row(ws, 2, len(headers))
    ws.freeze_panes = "A3"

    activations = [
        (1, "Position-jumping referral", "Day 1 (May 13)", "LIVE",
         "Claude",
         "Every signup gets pick1.live/?r=<code>. +100 positions per ref. 3 / 10 / 25 reward tiers."),
        (2, "Public Accountability Bet", "Day 8 (May 20)", "Planned",
         "Noa",
         "'If locks lose at < 60% over 30 days, first paid month refunded.' Needs public dashboard."),
        (3, "Beat the AI challenge", "Day 18 (May 30)", "Planned",
         "Claude (build) + Noa (launch)",
         "Twitter hashtag #BeatPick1. Weekly leaderboard. Winner gets year of Pro."),
        (4, "Influencer trade program", "Day 7 (May 19)", "Planned",
         "Noa (DMs) + Claude (list)",
         "DM 20 micro-influencers (5K-100K). Full list + DM template in launch/INFLUENCER_OUTREACH.md. Free Lifetime Pro + custom ref code. Anti-spam rules included."),
        (5, "Founder story drop", "Day 14-15 (May 26-27)", "Planned",
         "Noa (record) + Claude (cross-post)",
         "Coordinated multi-platform 'how I built this' content. Twitter thread + TikTok + IG + Substack."),
        (6, "Gleam Sweepstakes (FIFA WC Tickets + Lifetime Pro)", "Day 1 (May 13)", "LIVE",
         "Claude (setup) + Noa (prize fulfillment)",
         "✅ LIVE. URL: gleam.io/Ivb0j/win-2-fifa-world-cup-tickets-lifetime-pick1-pro. Title + description conversion-optimized May 14 (prize-first, $1,499 value). Free tier blockers documented (Feature Media, Viral Share, Repeat Action — all Hobby+). Runs through May 31, 11:59 PM ET. Drawn live."),
        (7, "SparkLoop Newsletter Cross-Promo", "Day 8 (May 20)", "Planned",
         "Claude (apply) + auto",
         "Pay $1-2/signup for Pick1 to appear in Morning Brew / The Hustle / sports newsletters' welcome emails. Budget cap: $200 for the 18-day window. Higher LTV than Meta cold. 24-72hr approval."),
        (8, "Live Waitlist Counter on pick1.live", "Day 2 (May 14)", "LIVE",
         "Claude",
         "✅ LIVE. Public Supabase RPC get_waitlist_count() returns greatest(real_count+1, 128). Frontend fetches via /api/waitlist-count (30s edge cache + 120s SWR). Replaces Telegram (dropped — crypto/tout adjacent)."),
        (9, "Sweeps CTA in email surfaces (welcome + daily pick)", "Day 2 (May 14)", "LIVE",
         "Claude",
         "✅ LIVE. Branded sweeps CTA block embedded in (a) welcome email - every new signup sees it under the 'daily pick' panel, (b) daily-pick email - every existing subscriber sees it every morning until May 31. UTM-tagged for attribution. ~17 days of repeat exposure to existing list."),
        (10, "18-channel distribution queue (Reddit / IH / Show HN / LinkedIn)", "Day 2 (May 14)", "Drafted",
         "Claude (drafts) + Noa (post)",
         "Paste-ready posts for r/SideProject (400K), r/giveaways (1M), r/contests (100K), r/free (1M), r/sweepstakes (45K), Indie Hackers milestone, LinkedIn founder, Hacker News Show HN (Day 18 Tuesday 8am ET). See launch/DISTRIBUTION_QUEUE.md + launch/TONIGHT.md. Account-gated — Noa posts each."),
        (11, "KingSumo viral giveaway (week 2)", "Day 10 (May 22)", "Planned",
         "Claude (setup) + Noa (prize)",
         "Free-tier KingSumo, 1,000-entry cap. Prize: 10× Lifetime Pro accounts ($0 in-kind cost, $9,990 perceived value). Built-in exponential referral multiplier - each entrant's odds scale with each friend they refer. Run in parallel with Gleam."),
        (12, "Zealy quest 'Pick1 Sprint'", "Day 2 (May 14)", "Planned",
         "Claude (spec) + Noa (account)",
         "1M+ quest hunters on Zealy. 10-task sprint with waitlist gateway. XP-based leaderboard, top 50 win Pro accounts. Quest config spec ready in launch/GROWTH_ACTIVATIONS.md A1. Free. Expected: 5-50K quest completions → 1.5-15K waitlist signups."),
        (13, "Galxe sports quest", "Day 2 (May 14)", "Planned",
         "Claude (spec) + Noa (account)",
         "13M+ users — largest quest platform. Same task list as Zealy. Off-chain only (no token integration needed). Free tier. Expected: 10-100K participants → 2-25K waitlist signups."),
        (14, "Layer3 quest", "Day 3 (May 15)", "Planned",
         "Noa",
         "1M+ users, higher-quality crypto-native audience. Free. Cross-list of same task structure. Expected: 2-20K participants → 500-5K signups."),
        (15, "TaskOn quest (APAC reach)", "Day 3 (May 15)", "Planned",
         "Noa",
         "3M+ users, APAC-dominant. Adds geographic spread + raw count. Free. Expected: 3-30K participants → 600-6K signups. Lower US conversion but counts."),
        (16, "[DROPPED] Pick1 Discord server", "—", "Cancelled",
         "—",
         "Cancelled May 14 — same maintenance burden as Telegram (which was dropped Day 1). Functions replaced by: (a) email 15 min before public post for 'early access' feel, (b) Twitter replies + IG comments for community, (c) Founder Pass for member-only status. Net signup impact ≈ 0 because the same users are still reachable via the other 21 activations."),
        (17, "Pick1 Founder Pass (digital collectible)", "Day 4 (May 16)", "Planned",
         "Claude (PNG generator) + Noa (creative approval)",
         "First 1,000 waitlist signups get a numbered Pick1 Founder Pass PNG (optionally minted free on Base L2 for crypto users; ~$10 total gas). Scarcity → 30-50% conversion lift on the signup form. Sent via Resend with unique serial."),
        (18, "Prediction-market community seeding", "Day 5-12", "Planned",
         "Noa",
         "Polymarket + Kalshi + Overtime Discord servers. Founder joins, becomes known commenter, posts high-value 7-day audit. Mentions Pick1 in passing — no shill. Expected 500-3K targeted signups."),
        (19, "Crypto sports Twitter deal (10-20 accounts)", "Day 3-5", "Planned",
         "Noa",
         "Pay 10-20 crypto-sports Twitter accounts (5-50K followers, >2% engagement) $50-100 USDC OR Lifetime Pro to retweet sweeps + 1 thread on Pick1. $500-2K cash or $0 in-kind. 200K-1M impressions, 500-5K signups."),
        (20, "TikTok #BeatPick1 challenge", "Day 6 (May 18)", "Planned",
         "Noa (launch) + founder team (engage)",
         "Daily AI pick reveal video — users duet with their own contrarian pick. Use trending sound. Founder team responds to every duet under 50 likes for algo boost. Expected 50-500 duets, 1-8K signups."),
        (21, "IG Story takeover program", "Day 7 (May 19)", "Planned",
         "Noa (recruit) + creators (post)",
         "10 micro creators (50-200K followers) each do 24-hr IG Story takeover on @pick1.live. Free in exchange for Lifetime Pro. Pre-vetted, founder-approved. 5-15K new IG followers."),
        (22, "Daily lineup card downloadable", "Day 10 (May 22)", "Planned",
         "Claude (build) + Noa (design)",
         "Beautifully designed daily card with Pick1's pick + 4 top games. Shareable IG Stories format. Embedded pick1.live CTA. Expected 10-30% of email subscribers share daily."),
        # ── $25K BUDGET ACTIVATIONS (added May 14 after budget authorization) ──
        (23, "🎯 Gleam Hobby tier upgrade", "Day 2 (May 14)", "Planned",
         "Noa (pay $39) + Claude (configure)",
         "$39/mo. Unlocks Feature Media (upload hero image), Viral Share entry method (multiplicative referrals), Repeat Action Limit (daily-redo social entries). Removes ALL Free-tier blockers we hit Day 2. Single highest-ROI move at $25K budget — first action."),
        (24, "💰 Meta Ads at scale", "Day 3 (May 15)", "Planned",
         "Noa (auth payment) + Claude (creative briefs)",
         "$500/day combined ($350 cold + $147 retargeting × 17 days = $8,500 total). At $1.50 CPL = 5-6K signups. Existing Campaign 1 already set up at $50/day — just scale the budget."),
        (25, "🎬 TikTok Spark Ads (reactive)", "Day 3 (May 15)", "Planned",
         "Noa (set up TikTok Ads Manager) + Claude (boost decisions)",
         "$175/day budget, reactive. Only boost Reels that organically pass 50K views. $3K cap over campaign. Decision window: within 6 hours of crossing threshold."),
        (26, "📣 1 × mid-tier sports creator", "Day 9 (May 21)", "Planned",
         "Noa (negotiate) + creator (post)",
         "$3K paid post from a 300-500K-follower sports/betting creator. Vet engagement >2% before signing. Outreach Day 4, contract Day 5, post drops Day 9. Single biggest single-shot reach line in $25K plan."),
        (27, "👥 5-7 × micro creators (paid)", "Day 7-12", "Planned",
         "Noa (outreach) + creators (post)",
         "$300-500 each, mix paid + product trade. From INFLUENCER_OUTREACH.md list, prioritized by engagement rate. Total: $2,500. DMs Day 7, posts roll Day 9-12."),
        (28, "✏️ Upwork content editor", "Day 5 (May 17)", "Planned",
         "Noa (hire) + editor (post)",
         "$500 one-off for 18-day gig. Daily cross-posting + basic editing while founder records. Frees founder time. Filipino or LatAm editor for ~$300; US-based $500-700."),
        (29, "🎞️ Production shoot day (street TikTok)", "Day 10 (May 22)", "Planned",
         "Noa (book) + creator (shoot)",
         "$1.5K one-day shoot in NYC or LA. Pro creator + lighting + edit. Produces 4-6 polished street-TikTok pieces. Block 2 backup weather dates."),
        (30, "💸 Sweepstakes prize-stack bump", "Day 12 (May 24)", "Planned",
         "Noa (announce + fund)",
         "$2,000 cash 2nd prize added to the WC tix grand prize. Mid-campaign announcement creates fresh news hook. Drives entry surge."),
        (31, "🚀 ProductHunt paid hunter", "Day 18 (May 30)", "Planned",
         "Noa (book) + hunter (launch day)",
         "$500 for an experienced hunter to do the launch coordination + animated demo + day-of comms on Day 21 (originally Day 21, accelerated to Day 18 with budget). Critical for top-3 daily."),
        (32, "📰 Press kit + media training", "Day 14 (May 26)", "Planned",
         "Noa (book session)",
         "$500 for professional photography + founder bio + 30-min media-training call. Sharpens press intake for the founder story drop wave."),
        (33, "🛠️ Infra upgrades (Resend Pro etc.)", "Day 5 (May 17)", "Planned",
         "Noa (pay) + Claude (configure)",
         "$200 one-off. Resend Pro (10K+ recipients), Vercel Pro (zero rate limit), Supabase Pro (production-tier). Needed once waitlist crosses 5K."),
    ]
    for r, row in enumerate(activations, start=3):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            if c == 6: cell.alignment = WRAP
        ws.row_dimensions[r].height = 38

    # Status colour coding — expanded range for 33 activations
    ws.conditional_formatting.add("D3:D40",
        CellIsRule(operator="equal", formula=['"LIVE"'],
                   fill=PatternFill("solid", fgColor="C6EFCE")))
    ws.conditional_formatting.add("D3:D40",
        CellIsRule(operator="equal", formula=['"Planned"'],
                   fill=PatternFill("solid", fgColor="FFEB9C")))
    ws.conditional_formatting.add("D3:D40",
        CellIsRule(operator="equal", formula=['"Drafted"'],
                   fill=PatternFill("solid", fgColor="DDEBF7")))
    ws.conditional_formatting.add("D3:D40",
        CellIsRule(operator="equal", formula=['"Blocked"'],
                   fill=PatternFill("solid", fgColor="FFC7CE")))
    ws.conditional_formatting.add("D3:D40",
        CellIsRule(operator="equal", formula=['"Cancelled"'],
                   fill=PatternFill("solid", fgColor="E0E0E0")))

    autosize(ws, [5, 36, 22, 11, 28, 65])


# ── Tab 6: Influencer Outreach ───────────────────────────────────────────
def build_influencer_outreach(wb):
    ws = wb.create_sheet("6 · Influencer Outreach")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    ws["A1"] = "INFLUENCER OUTREACH — 20-DM micro-trade program"
    ws["A1"].font = SECTION_FONT; ws["A1"].fill = SECTION_FILL; ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 32

    headers = [
        "Name / Handle", "Platform", "Followers", "Niche",
        "Status", "DM sent date", "Response", "Posted?",
        "Their ref code", "Notes",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=2, column=c, value=h)
    style_header_row(ws, 2, len(headers))
    ws.freeze_panes = "A3"

    # Pre-seed 20 blank rows for the trade program
    for i in range(20):
        ws.cell(row=3 + i, column=1, value=f"(slot {i+1})")
        ws.cell(row=3 + i, column=5, value="Cold")

    # Status colour coding
    ws.conditional_formatting.add("E3:E50",
        CellIsRule(operator="equal", formula=['"Cold"'],
                   fill=PatternFill("solid", fgColor="DDDDDD")))
    ws.conditional_formatting.add("E3:E50",
        CellIsRule(operator="equal", formula=['"DM sent"'],
                   fill=PatternFill("solid", fgColor="DDEBF7")))
    ws.conditional_formatting.add("E3:E50",
        CellIsRule(operator="equal", formula=['"Responded"'],
                   fill=PatternFill("solid", fgColor="FFEB9C")))
    ws.conditional_formatting.add("E3:E50",
        CellIsRule(operator="equal", formula=['"Posted"'],
                   fill=PatternFill("solid", fgColor="C6EFCE")))
    ws.conditional_formatting.add("E3:E50",
        CellIsRule(operator="equal", formula=['"Declined"'],
                   fill=PatternFill("solid", fgColor="FFC7CE")))

    autosize(ws, [24, 11, 11, 22, 11, 13, 26, 9, 17, 30])


# ── Tab 7: Pick Performance ──────────────────────────────────────────────
def build_pick_performance(wb):
    ws = wb.create_sheet("7 · Pick Performance")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:I1")
    ws["A1"] = "PICK PERFORMANCE — daily record (the receipts wall)"
    ws["A1"].font = SECTION_FONT; ws["A1"].fill = SECTION_FILL; ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 32

    headers = [
        "Date", "League", "Game", "Pick", "Probability",
        "Tier", "Result", "1u P&L", "Notes",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=2, column=c, value=h)
    style_header_row(ws, 2, len(headers))
    ws.freeze_panes = "A3"

    # Seed yesterday + today's picks (what's already in Supabase)
    picks = [
        (date(2026, 5, 12), "NBA", "SAS vs MIN G5",   "San Antonio Spurs",      78.6, "**",  "WIN",     "+0.91", "Series 3-2 SAS"),
        (date(2026, 5, 12), "NHL", "VGK @ ANA G5",    "Vegas Golden Knights",   57.7, "*",   "WIN",     "+0.91", "Tied series winner"),
        (date(2026, 5, 12), "NHL", "MTL vs BUF G4",   "Montreal Canadiens",     56.1, "*",   "LOSS",    "-1.00", "Missed G4"),
        (date(2026, 5, 13), "EPL", "MCI vs CRY",      "Manchester City",        81.2, "***", "Pending", "",      "Lock of the night"),
        (date(2026, 5, 13), "NHL", "COL @ MIN G5",    "Colorado Avalanche",     63.6, "**",  "Pending", "",      "Series closeout"),
        (date(2026, 5, 13), "NBA", "DET vs CLE G5",   "Detroit Pistons",        60.6, "*",   "Pending", "",      "Tied series G5"),
        (date(2026, 5, 13), "NHL", "MTL vs BUF G4",   "Montreal Canadiens",     56.1, "*",   "Pending", "",      "Carryover"),
        (date(2026, 5, 13), "NHL", "VGK @ ANA G5",    "Vegas Golden Knights",   57.7, "*",   "Pending", "",      "Carryover"),
    ]
    for r, p in enumerate(picks, start=3):
        for c, v in enumerate(p, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            if c == 1: cell.number_format = "yyyy-mm-dd"
            if c == 5: cell.number_format = "0.0%"

    # Result coloring
    ws.conditional_formatting.add("G3:G500",
        CellIsRule(operator="equal", formula=['"WIN"'],
                   fill=PatternFill("solid", fgColor="C6EFCE")))
    ws.conditional_formatting.add("G3:G500",
        CellIsRule(operator="equal", formula=['"LOSS"'],
                   fill=PatternFill("solid", fgColor="FFC7CE")))
    ws.conditional_formatting.add("G3:G500",
        CellIsRule(operator="equal", formula=['"Pending"'],
                   fill=PatternFill("solid", fgColor="DDDDDD")))

    autosize(ws, [12, 9, 22, 26, 11, 7, 11, 11, 28])

    # Summary row at the top: cumulative record
    ws["K2"] = "Summary"
    ws["K2"].font = HEADER_FONT
    ws["K2"].fill = HEADER_FILL
    ws["K3"] = "Total wins"
    ws["L3"] = '=COUNTIF(G:G,"WIN")'
    ws["K4"] = "Total losses"
    ws["L4"] = '=COUNTIF(G:G,"LOSS")'
    ws["K5"] = "Win %"
    ws["L5"] = '=IFERROR(L3/(L3+L4),"")'
    ws["L5"].number_format = "0.0%"
    ws["K6"] = "Lock (⭐⭐⭐) record"
    ws["L6"] = '=COUNTIFS(F:F,"***",G:G,"WIN")&"-"&COUNTIFS(F:F,"***",G:G,"LOSS")'
    ws.column_dimensions["K"].width = 24
    ws.column_dimensions["L"].width = 14


# ── Tab 8: $100K Scenario Budget ─────────────────────────────────────────
# A "what if we had 100x the budget" plan, for context. Real $1K plan is
# in MASTER_STRATEGY §10. This tab shows what scaling the same playbook
# would look like — useful when raising or pitching for marketing spend.
# Three columns: current $1K plan, $25K scenario (early traction), $100K
# scenario (post-seed marketing push). All figures are realistic
# market-rate estimates for the 18-day window.
def build_scenario_budget(wb):
    ws = wb.create_sheet("8 · $100K Scenario")
    ws.sheet_view.showGridLines = False

    # Section banner
    ws.merge_cells("A1:F1")
    ws["A1"] = "MARKETING BUDGET SCENARIOS — $1K (now) / $25K / $100K"
    ws["A1"].font = SECTION_FONT; ws["A1"].fill = SECTION_FILL; ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 32

    # Intro context block
    ws.merge_cells("A2:F2")
    ws["A2"] = ("This tab models what the same 18-day playbook would deliver at 25× and 100× the current $1,000 budget. "
                "All numbers are realistic 2026 market rates. Use this when pitching seed/series-A on what marketing dollars buy.")
    ws["A2"].font = SUBTLE_FONT
    ws["A2"].alignment = WRAP
    ws.row_dimensions[2].height = 36

    # Headers
    headers = ["Line item", "What it gets you", "$1K (current)", "$25K (early traction)", "$100K (post-seed)", "Notes"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=4, column=c, value=h)
    style_header_row(ws, 4, len(headers))
    ws.freeze_panes = "A5"

    rows = [
        # PAID ACQUISITION
        ("[Paid Acquisition]", "", "", "", "", ""),
        ("Meta Ads — cold", "Lookalikes + interest stacks", 400, 6000, 15000,
         "$1.50 CPL target. $400 = 250 signups; $6K = 4K signups; $15K = 10K signups."),
        ("Meta Ads — retargeting", "Pixel-visited but didn't sign up", 150, 2500, 5000,
         "Always higher CTR than cold. 3-7x lower CPL. Burns hot."),
        ("TikTok Spark Ads", "Boost organic Reels that pop", 150, 3000, 10000,
         "Only deploy on Reels that already crossed 50K views organically. Otherwise it's lighting money on fire."),
        # ── Subtotal row
        ("  → Subtotal: paid digital", "", "=SUM(C6:C8)", "=SUM(D6:D8)", "=SUM(E6:E8)", ""),

        # INFLUENCER
        ("[Influencer]", "", "", "", "", ""),
        ("Macro creator (1, 1M+ followers)", "Sports/betting niche, 1 sponsored post or video w/ exclusive code", 0, 0, 20000,
         "$15-25K market rate for a single post from a 1M+ sports creator. Single biggest line item at $100K — but the lift is exponential."),
        ("Mid-tier creators (4-6, 200-500K)", "Sponsored IG/TikTok posts, mix paid + product-trade", 0, 5000, 15000,
         "$2-3K avg per post for 200-500K followers. Best CAC/LTV mix."),
        ("Micro creators (15-20, 50-150K)", "Paid + product-trade mix; ref-code attribution", 0, 2000, 10000,
         "$500-1K per post. Highest engagement rate; 20 small bets beat 1 big one."),
        ("Influencer trades (free Pro)", "$0 cost — Lifetime Pro for honest review (10-20 micros)", 0, 0, 0,
         "Always-on. $0 line item but real value. 20-creator DM list already in tab 6."),
        ("  → Subtotal: influencer", "", "=SUM(C11:C14)", "=SUM(D11:D14)", "=SUM(E11:E14)", ""),

        # PRODUCTION / CONTENT
        ("[Content Production]", "", "", "", "", ""),
        ("Live 'man on the street' TikTok shoots", "Multi-city sports-fan reactions to AI picks", 0, 4000, 18000,
         "Pro creator + videographer + 4-8 city shoots ($1.5-2.5K each) + edits + music licensing. Closest thing to 'guaranteed viral' for sports."),
        ("Branded creative (Canva Pro Team, Pond5, motion design)", "Evergreen content library", 50, 1000, 3000,
         "$50 currently covers Canva Pro + Pond5 single licenses. $3K = pro motion designer for 4-6 hero animations."),
        ("Studio/voiceover for hero ad", "Single high-production-value hero ad", 0, 0, 5000,
         "Skip at $1K and $25K. At $100K it's worth it for the YouTube pre-roll + paid social hero asset."),
        ("  → Subtotal: production", "", "=SUM(C17:C19)", "=SUM(D17:D19)", "=SUM(E17:E19)", ""),

        # SWEEPS / VIRAL
        ("[Sweeps & Viral Mechanics]", "", "", "", "", ""),
        ("Sweepstakes prize budget", "Headline prize value", 500, 5000, 25000,
         "Currently: 2 WC tickets ($500). $25K scenario: upgrade to $5K cash. $100K scenario: $25K cash or 5 winners × $5K each — drives 10x more entries."),
        ("Gleam Hobby/Pro plan", "Unlocks Feature Media, Viral Share, Repeat Action", 0, 39, 200,
         "$39/mo Hobby unlocks Feature Media + Repeat Action. $200/mo Business unlocks Custom CSS + advanced anti-fraud. Currently blocked us tonight."),
        ("KingSumo / Woorise paid tier", "Higher entry caps + advanced anti-bot", 0, 0, 500,
         "$500 unlocks 10K+ entries cap, custom branding, removes 'Powered by' footer. Worth at scale."),
        ("SparkLoop newsletter cross-promo", "$1-2 / signup via Morning Brew, The Hustle, etc.", 0, 2000, 8000,
         "100-200 signups at $200 budget. Scale linearly. Higher LTV than Meta cold."),
        ("Daily Drop on Twitter (7×$25 prizes)", "Daily micro-sweeps the week before launch", 0, 175, 700,
         "$25/day × 7 days = $175 currently. At $100K bump to $100/day × 7 = $700."),
        ("  → Subtotal: sweeps & viral", "", "=SUM(C22:C26)", "=SUM(D22:D26)", "=SUM(E22:E26)", ""),

        # PR & EARNED MEDIA
        ("[PR & Earned Media]", "", "", "", "", ""),
        ("PR firm retainer (1 month, mid-tier)", "Targeted outreach to The Hustle, Morning Brew, sports media", 0, 0, 3000,
         "1 month with a small/mid-tier firm. Worth it if even one Hustle/MB pickup lands."),
        ("Press kit + media training", "Photography, founder bio, talking points", 0, 500, 1500,
         "Professional shots + crisp positioning for press intake."),
        ("ProductHunt coordinated launch", "Hunter, day-of comms, video, asset polish", 0, 0, 1500,
         "Free at $1K (DIY). $1.5K covers a paid hunter + animated demo for PH listing."),
        ("Hacker News Show HN", "Free — just need polish + timing", 0, 0, 0,
         "$0 line. Timing-sensitive (Tuesday 8am ET). Front page = 50K-200K visits."),
        ("  → Subtotal: PR & earned", "", "=SUM(C29:C32)", "=SUM(D29:D32)", "=SUM(E29:E32)", ""),

        # OTHER / TOOLING
        ("[Tooling & Infra]", "", "", "", "", ""),
        ("Resend / Supabase / Vercel / etc.", "Email + DB + hosting", 0, 100, 500,
         "Free tier suffices at $1K-25K. $500/mo plans needed at scale (Resend Pro, Supabase Pro, Vercel Pro)."),
        ("Analytics / attribution stack", "PostHog, Plausible, GA4 setup", 0, 100, 1000,
         "DIY at $1K (free tiers). $1K covers a 1-time analytics consultant to build proper funnels."),
        ("Founder time (Noa) — opportunity cost", "Not a line item, but real", 0, 0, 0,
         "Excluded from this budget. Worth noting."),
        ("  → Subtotal: tooling", "", "=SUM(C35:C37)", "=SUM(D35:D37)", "=SUM(E35:E37)", ""),

        # CONTINGENCY
        ("[Contingency]", "", "", "", "", ""),
        ("Reallocation reserve", "Day 10 reallocate to top-performing channel", 250, 1500, 5000,
         "Critical. Whichever channel is winning by Day 10, dump 30% of the budget into it."),

        # ── TOTAL
        ("", "", "", "", "", ""),
        ("TOTAL", "", "=C6+C7+C8+C11+C12+C13+C14+C17+C18+C19+C22+C23+C24+C25+C26+C29+C30+C31+C32+C35+C36+C37+C41",
                     "=D6+D7+D8+D11+D12+D13+D14+D17+D18+D19+D22+D23+D24+D25+D26+D29+D30+D31+D32+D35+D36+D37+D41",
                     "=E6+E7+E8+E11+E12+E13+E14+E17+E18+E19+E22+E23+E24+E25+E26+E29+E30+E31+E32+E35+E36+E37+E41",
                     "Excluding $0 in-kind items + founder time"),
    ]

    # Find total row index
    for r, row in enumerate(rows, start=5):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = WRAP if c in (2, 6) else (CENTER if c >= 3 and c <= 5 else LEFT)
            # Section banners (rows where col B is empty and col A starts with [)
            if str(row[0]).startswith("[") and str(row[1]) == "":
                cell.font = Font(name="Calibri", bold=True, size=11, color=BRAND_BG)
                cell.fill = PatternFill("solid", fgColor="F4F4F4")
            # Subtotal rows
            elif str(row[0]).strip().startswith("→"):
                cell.font = Font(name="Calibri", bold=True, size=10, color="333333")
                cell.fill = PatternFill("solid", fgColor="EFEFEF")
            # TOTAL row
            elif row[0] == "TOTAL":
                cell.font = Font(name="Calibri", bold=True, size=12, color=BRAND_BG)
                cell.fill = HEADER_FILL
            # Number cols
            if c in (3, 4, 5) and isinstance(v, (int, float)):
                cell.number_format = "$#,##0"
            elif c in (3, 4, 5) and isinstance(v, str) and v.startswith("="):
                cell.number_format = "$#,##0"
        ws.row_dimensions[r + 0].height = 30

    autosize(ws, [40, 50, 14, 18, 18, 60])

    # ── Expected outcomes section ──
    out_row = len(rows) + 7  # leave a gap

    ws.merge_cells(f"A{out_row}:F{out_row}")
    ws.cell(row=out_row, column=1, value="EXPECTED OUTCOMES (18-day pre-launch sprint)")
    ws.cell(row=out_row, column=1).font = SECTION_FONT
    ws.cell(row=out_row, column=1).fill = SECTION_FILL
    ws.cell(row=out_row, column=1).alignment = CENTER
    ws.row_dimensions[out_row].height = 32

    outcome_headers = ["Metric", "Target", "$1K plan", "$25K plan", "$100K plan", "Notes"]
    for c, h in enumerate(outcome_headers, start=1):
        ws.cell(row=out_row + 1, column=c, value=h)
    style_header_row(ws, out_row + 1, len(outcome_headers))

    outcomes = [
        ("Waitlist signups (May 31)", "1,500", "1,500-2,500", "8,000-15,000", "25,000-50,000",
         "Linear scaling on Meta + influencer; sub-linear on organic (organic has reach ceiling)."),
        ("Cost per waitlist signup (blended)", "<$1.50", "$0.40-$0.67", "$1.67-$3.13", "$2.00-$4.00",
         "Goes UP with scale (you can't fake-organic at scale). Still cheap vs SaaS norms ($25-100 CAC)."),
        ("Instagram followers gained", "—", "500-1.5K", "5K-15K", "20K-50K",
         "Influencer + paid TikTok cross-pollinate IG."),
        ("TikTok followers gained", "—", "1K-3K", "8K-20K", "30K-75K",
         "TikTok rewards consistency + paid Spark Ads. Biggest multiplier at $100K."),
        ("Twitter followers gained", "—", "200-500", "2K-5K", "8K-20K",
         "Twitter is hardest to grow organically. PR + Show HN moves the needle here."),
        ("Press placements", "1-2", "1 small", "3-5 mid", "5-10 incl. 1-2 tier-1",
         "$1K = founder hustle. $25K = niche placements (Sports Illustrated digital, Bleacher Report). $100K = potential TechCrunch / The Hustle."),
        ("Viral pieces (>100K views, organic)", "1", "0-1", "2-4", "4-8",
         "Need quality + volume. Live street TikToks at $100K specifically engineered for this."),
        ("Brand recall in target audience (post-launch survey)", "—", "<5%", "10-20%", "25-40%",
         "Estimated via post-launch user survey. At $100K + a press hit, brand becomes a known name in the bettor demo."),
        ("Lifetime Pro conversions (% of signups)", "1-2%", "15-30 ($45-450 ARR)", "80-300 ($240-9K ARR)", "250-1.5K ($750-45K ARR)",
         "Conservative. Lifetime Pro at $30 launch promo, $150 regular. Actual conversion depends on launch-week comms."),
    ]
    for r, row in enumerate(outcomes, start=out_row + 2):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = WRAP if c in (2, 3, 4, 5, 6) else LEFT
            cell.font = BODY_FONT
        ws.row_dimensions[r].height = 36

    # ── Strategy notes section ──
    notes_row = out_row + 2 + len(outcomes) + 2

    ws.merge_cells(f"A{notes_row}:F{notes_row}")
    ws.cell(row=notes_row, column=1, value="STRATEGY NOTES — what changes at each scale")
    ws.cell(row=notes_row, column=1).font = SECTION_FONT
    ws.cell(row=notes_row, column=1).fill = SECTION_FILL
    ws.cell(row=notes_row, column=1).alignment = CENTER
    ws.row_dimensions[notes_row].height = 32

    notes = [
        ("$1K (current):",
         "Founder-led hustle. Every dollar of ad spend is rationed. Distribution is paste-and-post + organic + sweeps. Goal is to PROVE the funnel works before raising. Target CPL <$1.50, founder time covers everything else."),
        ("$25K (early traction):",
         "First serious paid push. Hire 1-3 micro creators for paid posts. Upgrade Gleam to Hobby to unlock branded image + viral share. Add a small PR push (DIY founder outreach with media training). 5-10x the signup count of $1K plan."),
        ("$100K (post-seed marketing):",
         "Real campaign. Macro influencer ($15-25K single post), multi-city 'man on the street' TikTok shoots ($18K), bigger sweeps prize ($25K cash gets 10x more entrants than $500 tickets), PR firm retainer, ProductHunt-coordinated launch with paid hunter. Brand starts to feel like a 'thing' in the sports-bettor demo."),
        ("Diminishing returns above $100K:",
         "Once you're at 50K+ pre-launch waitlist, the bottleneck becomes product, not marketing. Spending more on awareness without conversion infrastructure (onboarding, retention loop, paid tier) is wasted. Set the marketing pause at signup target."),
    ]
    for r, (label, txt) in enumerate(notes, start=notes_row + 1):
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=11, color=BRAND_BG)
        ws.merge_cells(f"B{r}:F{r}")
        ws.cell(row=r, column=2, value=txt)
        ws.cell(row=r, column=2).font = BODY_FONT
        ws.cell(row=r, column=2).alignment = WRAP
        ws.row_dimensions[r].height = 52


# ── Tab 9: Daily Strategy ────────────────────────────────────────────────
# The single most important operational tab. One row per day from today
# (May 14, Day 2) through launch (May 31, Day 19). For each day: theme,
# what's auto-running (Claude's cron + built-ins), what the user has to
# post (paste-ready, per platform), outreach actions, activations
# launching, daily signup target, and cumulative target. Color-coded by
# campaign phase (foundation / amplify / final push).
def build_daily_strategy(wb):
    ws = wb.create_sheet("9 · Daily Strategy")
    ws.sheet_view.showGridLines = False

    # Section banner
    ws.merge_cells("A1:I1")
    ws["A1"] = "DAILY STRATEGY — 18-day execution plan (May 14 → May 31)"
    ws["A1"].font = SECTION_FONT; ws["A1"].fill = SECTION_FILL; ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 32

    # Intro
    ws.merge_cells("A2:I2")
    ws["A2"] = ("THE source-of-truth daily plan. Goal: maximize waitlist signups + social followers. "
                "Auto column = Claude's cron + autonomous infra. You column = paste-ready posts you ship. "
                "Outreach = personal asks / DMs. Activation column = what launches or runs that day. "
                "Cumulative target = where the waitlist needs to be by EOD.")
    ws["A2"].font = SUBTLE_FONT; ws["A2"].alignment = WRAP
    ws.row_dimensions[2].height = 44

    # Headers
    headers = ["Date", "Day", "Focus", "Auto (Claude)",
               "You — posts to ship", "You — outreach / setup",
               "Activations live/launching", "Daily target", "Cumulative"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=4, column=c, value=h)
    style_header_row(ws, 4, len(headers))
    ws.freeze_panes = "C5"  # freeze first 2 cols + header rows

    # Phase color fills
    PHASE_FOUNDATION = PatternFill("solid", fgColor="E3F2FD")  # light blue
    PHASE_AMPLIFY    = PatternFill("solid", fgColor="E8F5E9")  # light green
    PHASE_FINAL      = PatternFill("solid", fgColor="FFF3E0")  # light orange
    PHASE_LAUNCH     = PatternFill("solid", fgColor="FFE0B2")  # darker orange

    # daily rows: (date, day#, focus, auto, posts, outreach, activations, target, cum)
    days = [
        # ── WEEK 1: FOUNDATION (Day 2-7, May 14-19) ──
        ("May 14", 2, "🚀 SWEEPS LAUNCH NIGHT",
         "• Sweeps CTA live in welcome + daily-pick emails\n• pick1.live banner updated with new URL",
         "• X: sweeps announcement (SWEEPSTAKES_LAUNCH_COPY.md §1)\n• IG: feed post + update bio link (§2)\n• FB Page: post + pin (§4)\n• Email: blast to existing waitlist (§5)\n• LinkedIn: founder post (DISTRIBUTION_QUEUE.md)\n• Reddit r/SideProject (DISTRIBUTION_QUEUE.md)",
         "• Launch Meta Campaign 1 ($50/day, 3 PNGs)\n• Create Zealy + Galxe + Layer3 + TaskOn accounts (~20 min)\n• Send Twitter dev app creds (4 OAuth values)",
         "✅ Gleam LIVE (#6)\n✅ Email CTAs LIVE (#9)\n✅ Distribution queue drafted (#10)\nLaunching: Sweeps announcement everywhere",
         150, 250),
        ("May 15", 3, "🌐 WEB3 QUEST LAUNCH",
         "• Build Founder Pass PNG generator\n• Ship referral leaderboard page on pick1.live\n• Configure Zealy + Galxe + Layer3 + TaskOn quests (after Noa creates accounts)",
         "• X: 'thread on how Pick1 calibrates' (organic)\n• IG: carousel post (last 3 days picks)\n• Reddit r/giveaways post (DISTRIBUTION_QUEUE.md)\n• Reddit r/contests post\n• Indie Hackers milestone post",
         "• Crypto Twitter outreach: 5 accounts (5-50K followers) — $50-100 each or Lifetime Pro\n• Reach out to first 3 micro-influencers",
         "Launching: Zealy (#12), Galxe (#13), Layer3 (#14), TaskOn (#15) quests\nLaunching: Founder Pass (#17) — first batch",
         200, 450),
        ("May 16", 4, "🎬 CONTENT BLITZ",
         "• Send Founder Pass PNG to all existing waitlist via Resend\n• Render 2 new ad creative variations for A/B test",
         "• TikTok: founder explainer video (30 sec)\n• IG Reel: same content, vertical crop\n• X: quote-tweet yesterday's sweeps with thread\n• FB: share Founder Pass announcement",
         "• Crypto Twitter: 5 more accounts\n• DM 3 micro-influencers",
         "✅ Founder Pass live (#17)\nWeb3 quests trickling in (#12-15)",
         200, 650),
        ("May 17", 5, "🎯 TARGETED COMMUNITY SEEDING",
         "• Daily-tweet cron live (after Twitter dev creds)\n• Mid-week dashboard email to existing list",
         "• TikTok: man-on-the-street style pick reveal\n• IG Story: poll 'who's tonight's winner?'\n• X: quote ratio-bet against a bad take",
         "• Polymarket Discord: founder lurks/comments\n• Kalshi Discord: same\n• Overtime Markets: same\n• DM 3 more micro-influencers",
         "Launching: Prediction-market seeding (#18)\nRunning: Web3 quests, sweeps, ads",
         200, 850),
        ("May 18", 6, "📱 TIKTOK CHALLENGE LAUNCH",
         "• Daily pick auto-tweet\n• Auto sweeps reminder in daily-pick email (already live, ongoing)",
         "• TikTok: launch #BeatPick1 challenge with trending sound\n• IG: same video reposted as Reel\n• X: launch tweet for #BeatPick1\n• Respond to every TikTok duet under 50 likes",
         "• Crypto Twitter: 5 more accounts (total 15)\n• Polymarket Discord: post 7-day audit",
         "Launching: TikTok #BeatPick1 (#20)\nRunning: All quests, sweeps, ads, IG Story takeover #1 starts",
         250, 1100),
        ("May 19", 7, "📨 INFLUENCER OUTREACH DAY",
         "• Mid-campaign sweeps reminder email blast (auto, all subscribers)\n• Build Daily Lineup Card generator",
         "• Daily pick + tweet\n• IG Story: behind the scenes of how Pick1 picks\n• TikTok: respond to #BeatPick1 duets in a compilation video\n• Reddit r/free post",
         "• ★ DM all 20 micro-influencers from INFLUENCER_OUTREACH.md\n• Personalized first line is non-negotiable\n• IG Story takeover #2 launches",
         "Launching: Influencer program (#4) full DM blast\nLaunching: IG Story takeovers (#21) — day 1",
         300, 1400),
        # ── WEEK 2: AMPLIFY (Day 8-14, May 20-26) ──
        ("May 20", 8, "📰 NEWSLETTER CROSS-PROMO",
         "• Submit to SparkLoop ($200 cap) — 24-72hr approval\n• Public Accountability Bet page goes live",
         "• Daily pick + tweet\n• TikTok: 'AI bet vs human bet' format\n• IG: today's pick as branded card",
         "• IG Story takeover #3\n• Follow up with non-responding influencers (1 ask only, then drop)",
         "Launching: SparkLoop (#7)\nLaunching: Public Accountability Bet (#2)\nRunning: All quests, sweeps, influencer takeovers",
         300, 1700),
        ("May 21", 9, "🎨 CONTENT WEEKEND",
         "• Saturday — sports content peaks. Auto-send 'weekend slate' email to existing list",
         "• TikTok: weekend best-of compilation\n• IG Story: live game commentary as picks play out\n• X: thread on weekend sports betting biggest opportunities",
         "• IG Story takeover #4\n• Crypto Twitter posts going live (10-20 retweets accumulating)",
         "Running: All quests, ads, sweeps, content cadence",
         300, 2000),
        ("May 22", 10, "🏆 WEEK 2 LAUNCH STACK",
         "• Day-10 budget reallocation: kill ads >$3 CPL, double winners\n• Launch KingSumo viral giveaway (free tier, 1K cap)\n• Daily Lineup Card live in morning email",
         "• TikTok: 'AI got X right last week' receipts video\n• IG: carousel of last week's hits\n• X: thread of the week's wins + a miss (transparency)",
         "• IG Story takeover #5\n• 7-day check-in with first responding influencers",
         "Launching: KingSumo viral giveaway (#11)\nLaunching: Daily Lineup Card (#22)",
         400, 2400),
        ("May 23", 11, "🔁 RETARGETING WAVE",
         "• Bump Meta retargeting budget to $20/day (anyone who visited but didn't sign up)\n• Re-engagement email to 7-day-cold subscribers",
         "• Daily pick + tweet\n• IG: today's lineup card\n• TikTok: 'POV: you bet on Pick1' format",
         "• IG Story takeover #6\n• Quote-tweet last week's pick with the result",
         "Running: KingSumo + Gleam in parallel + all quests + IG takeovers",
         350, 2750),
        ("May 24", 12, "📊 RECEIPTS WEEKEND",
         "• Auto-send 'two-week receipts' email — every pick we've made + result",
         "• TikTok: 'I posted every pick before kickoff — here's the result' format\n• IG: 2-week receipts carousel\n• X: thread on calibration so far",
         "• IG Story takeover #7\n• Reddit r/sportsbook brief comment-engagement (not the big post yet)",
         "Running: All. Content density peaks for the weekend.",
         350, 3100),
        ("May 25", 13, "🎤 PRE-FOUNDER-STORY HYPE",
         "• Pre-warm email subscribers: 'tomorrow we tell you why Pick1 exists'",
         "• X: cryptic tease tweet\n• IG Story: BTS shot of the founder writing\n• TikTok: '24 hours until I tell you the real story'",
         "• IG Story takeover #8\n• Send press list a heads-up (1-line email): 'founder essay drops tomorrow'",
         "Running: All quests + sweeps + KingSumo. Building anticipation.",
         300, 3400),
        ("May 26", 14, "📖 FOUNDER STORY DROP - DAY 1",
         "• Auto-email blast: full founder essay to existing list\n• Auto-pin essay on pick1.live homepage",
         "• X: long-form thread (1,500 words, 25 tweets)\n• Substack/Medium: full 2,000-word essay\n• TikTok: 60-sec founder-to-camera version\n• IG carousel: essay highlights",
         "• DM essay link to 10 sports journalists with personal note\n• IG Story takeover #9",
         "Launching: Founder Story Drop (#5)\nBig content wave — set engagement alerts",
         600, 4000),
        # ── WEEK 3: FINAL PUSH (Day 15-19, May 27-31) ──
        ("May 27", 15, "🔄 FOUNDER STORY DROP - DAY 2 (AMPLIFY)",
         "• Auto-boost: pin best comments + quote-tweet best responses",
         "• X: respond to every reply on the thread (algo signal)\n• TikTok: 'reply to the founder story' video\n• IG: respond to DMs + Story replies",
         "• Push top-performing piece to paid amplification ($200 TikTok Spark Ads if Reel passed 50K)\n• IG Story takeover #10 (final)",
         "Boost mode — riding the founder story wave",
         500, 4500),
        ("May 28", 16, "🥊 BEAT THE AI CHALLENGE LAUNCH",
         "• Auto-launch /beat-the-ai page (Twitter hashtag tracker + leaderboard)",
         "• X: launch #BeatPick1 hashtag campaign\n• TikTok: founder challenge video\n• IG: explainer Reel",
         "• Twitter ratio-bet content drop (designed to be quote-dunked)\n• Reddit r/sportsbook: long-form analysis post (DISTRIBUTION_QUEUE.md style)",
         "Launching: Beat the AI Challenge (#3)\nReddit r/sportsbook post (heaviest non-paid activation)",
         500, 5000),
        ("May 29", 17, "📺 SHOW HN PREP + SOCIAL PROOF",
         "• Pre-stage Show HN draft + screenshots\n• Auto-collect best testimonials from existing users for a 'wall of love' page",
         "• Sports betting forum posts: covers.com, sportsbookreview, betpop\n• TikTok: 'wall of love' compilation\n• IG: testimonials carousel",
         "• Cross-promote with 1-2 prediction-market accounts for AMA",
         "Building social proof for HN + launch day",
         400, 5400),
        ("May 30", 18, "💥 SHOW HN DAY + DAILY DROP STARTS",
         "• 8:00 AM ET: Post Show HN (DISTRIBUTION_QUEUE.md draft)\n• Auto-respond to comments via founder, not bot\n• Daily Drop #1 on Twitter (cron, $25 prize)\n• Auto-blast 'last 24 hours' sweeps reminder to all signups",
         "• X: Pin HN link, thread it\n• TikTok: 'Pick1 is on the front page of HN' format if it hits\n• IG: launch-week countdown content",
         "• Pre-warm: schedule winner-draw livestream announcement\n• Send press kit to anyone covering HN",
         "Launching: Show HN, Daily Drop (#3 sub-mechanic), Final reminder blast",
         600, 6000),
        ("May 31", 19, "🏁 LAUNCH DAY + WINNER DRAW",
         "• Auto: final 6hr / 3hr / 1hr countdown emails\n• Auto: winner-draw stream landing page goes live\n• Auto: post-draw winner announcement email\n• Auto: 'app launches Monday' transition email",
         "• X: 'sweeps closing in 6 hours' urgency posts every 2 hours\n• IG Story: live countdown\n• TikTok: live draw if possible\n• FB Live: draw event 11:59 PM ET",
         "• Press: 'winner announced live' moment for screenshots\n• Daily Drop #2 + Daily Drop #3 (final 2 prizes)",
         "All activations close. Draw happens at 23:59 ET.",
         800, 6800),
    ]

    for r_idx, row in enumerate(days, start=5):
        date, day_n, focus, auto, posts, outreach, acts, target, cum = row
        # Phase color: Week 1 = foundation, Week 2 = amplify, Week 3 = final push, Day 19 = launch
        if day_n == 19:
            phase = PHASE_LAUNCH
        elif day_n >= 15:
            phase = PHASE_FINAL
        elif day_n >= 8:
            phase = PHASE_AMPLIFY
        else:
            phase = PHASE_FOUNDATION

        ws.cell(row=r_idx, column=1, value=date).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=r_idx, column=2, value=day_n).font = Font(name="Calibri", bold=True, size=11)
        ws.cell(row=r_idx, column=3, value=focus).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=r_idx, column=4, value=auto).font = BODY_FONT
        ws.cell(row=r_idx, column=5, value=posts).font = BODY_FONT
        ws.cell(row=r_idx, column=6, value=outreach).font = BODY_FONT
        ws.cell(row=r_idx, column=7, value=acts).font = BODY_FONT
        ws.cell(row=r_idx, column=8, value=target).number_format = "#,##0"
        ws.cell(row=r_idx, column=9, value=cum).number_format = "#,##0"

        # Apply phase fill to focus cell
        ws.cell(row=r_idx, column=3).fill = phase

        # Wrap text everywhere relevant
        for c in (3, 4, 5, 6, 7):
            ws.cell(row=r_idx, column=c).alignment = WRAP
        for c in (1, 2, 8, 9):
            ws.cell(row=r_idx, column=c).alignment = CENTER

        # Row height — taller for content-heavy days
        ws.row_dimensions[r_idx].height = 130

    autosize(ws, [11, 5, 22, 38, 42, 36, 38, 11, 12])

    # ── Strategy notes section ──
    notes_row = len(days) + 7

    ws.merge_cells(f"A{notes_row}:I{notes_row}")
    ws.cell(row=notes_row, column=1, value="HOW TO READ THIS TAB")
    ws.cell(row=notes_row, column=1).font = SECTION_FONT
    ws.cell(row=notes_row, column=1).fill = SECTION_FILL
    ws.cell(row=notes_row, column=1).alignment = CENTER
    ws.row_dimensions[notes_row].height = 32

    notes = [
        ("Phase colors",
         "🔵 Light blue = Week 1 Foundation (build the funnel, push launch). 🟢 Light green = Week 2 Amplify (compound through cross-promo + influencers). 🟠 Light orange = Week 3 Final Push (founder story → HN → launch). Each phase has a different rhythm — Week 1 ships hard, Week 2 nurtures, Week 3 hypes."),
        ("Daily rhythm (every day)",
         "✓ Morning (9am ET): daily pick auto-sends to email + tweet (cron). ✓ Mid-morning: you post the day's TikTok + IG content. ✓ Afternoon: respond to comments on yesterday's posts. ✓ Evening: 1-3 outreach DMs OR amplify a hot piece of content. ✓ Night: check tracker, update KPI Dashboard."),
        ("If Meta CPL > $3",
         "Day 10 reallocation rule (May 22): kill the worst ad, double the budget on the winner. Don't slowly drift — be ruthless. The Day 10 decision determines whether weeks 2-3 succeed."),
        ("If a TikTok hits 50K+ views organically",
         "Spend the $150 TikTok Spark Ads budget on it. Don't try to predict which one — let the algorithm tell you. Boost decision should be within 6 hours of crossing the threshold."),
        ("If the founder story drop bombs",
         "Don't double down. Move budget to whatever channel IS working (probably Meta retargeting + the winning TikTok). The story drop is a swing — has to be allowed to miss."),
        ("If Show HN flops",
         "Don't post again. Show HN gets one shot — post twice and you get downvote-bombed. Focus on Reddit r/sportsbook + the draw-day moment instead."),
        ("Cumulative targets",
         "Daily targets are conservative. Hitting them all = 6,800 signups by EOD May 31. Stretch (2x daily) = 13,600. Baseline goal is 1,500-2,500; everything above that is bonus from the web3 + influencer layers."),
        ("Track every day",
         "Tab 1 (KPI Dashboard) holds the metrics. This tab is the PLAN; tab 1 is the SCOREBOARD. Update tab 1 every night before bed (~3 min)."),
    ]
    for r, (label, txt) in enumerate(notes, start=notes_row + 1):
        ws.cell(row=r, column=1, value=label).font = Font(name="Calibri", bold=True, size=11, color=BRAND_BG)
        ws.merge_cells(f"B{r}:I{r}")
        ws.cell(row=r, column=2, value=txt).font = BODY_FONT
        ws.cell(row=r, column=2).alignment = WRAP
        ws.row_dimensions[r].height = 48


# ── Tab 10: Content Engine ───────────────────────────────────────────────
# Daily content production blueprint. Specifies what to post when on
# which platform, with repurposing matrix and hashtag library. Built as
# CMO would design it: aggressive content velocity + 6-8× repurposing
# multiplier so each piece of original content shows up everywhere.
def build_content_engine(wb):
    ws = wb.create_sheet("10 · Content Engine")
    ws.sheet_view.showGridLines = False

    # Section banner
    ws.merge_cells("A1:F1")
    ws["A1"] = "CONTENT ENGINE — daily production blueprint (22-34 touchpoints/day)"
    ws["A1"].font = SECTION_FONT; ws["A1"].fill = SECTION_FILL; ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 32

    # Intro
    ws.merge_cells("A2:F2")
    ws["A2"] = ("CMO call: with $1K budget you can't out-spend, only out-volume + out-repurpose. "
                "Current state: 2 TikToks/day. Target: 6 TikToks/day × 6-8x repurposing = 22-34 touchpoints/day. "
                "Founder time: ~70-80 min/day. Add a $300-500 Upwork content editor for the 18 days "
                "if 80 min/day is unsustainable.")
    ws["A2"].font = SUBTLE_FONT; ws["A2"].alignment = WRAP
    ws.row_dimensions[2].height = 44

    # ── Daily content schedule (24-hour blueprint) ──
    ws.merge_cells("A4:F4")
    ws["A4"] = "DAILY CONTENT SCHEDULE (run this every day)"
    ws["A4"].font = Font(name="Calibri", bold=True, size=12, color=BRAND_BG)
    ws["A4"].fill = PatternFill("solid", fgColor="EFEFEF")
    ws["A4"].alignment = LEFT
    ws.row_dimensions[4].height = 22

    headers = ["Time (ET)", "Platform", "Content type", "Effort", "Auto/Manual", "Notes"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=5, column=c, value=h)
    style_header_row(ws, 5, len(headers))

    schedule = [
        ("6:00 AM", "TikTok", "🎯 Pick-of-the-day reveal (45-sec founder selfie)", "5 min", "Manual", "Hook: 'Wait for it...' or 'POV: tonight's lock'"),
        ("7:00 AM", "Twitter cron", "Daily AI pick + confidence %", "0", "Auto", "Once dev creds wired in (Day 5)"),
        ("7:05 AM", "Instagram Reel", "Cross-post the 6am TikTok", "30 sec", "Manual", "Same vertical 1080×1920 file"),
        ("7:10 AM", "YouTube Short", "Cross-post the 6am TikTok", "30 sec", "Manual", "Same file. Free distribution."),
        ("7:15 AM", "IG Story", "Sticker poll: 'over or under?'", "1 min", "Manual", "Engagement signal for the algo"),
        ("8:00 AM", "Email cron", "Daily-pick email (with sweeps CTA)", "0", "Auto", "Already live in production"),
        ("9:00 AM", "Twitter", "Reply-bomb 5 sports accounts with Pick1's probability", "10 min", "Manual", "★ Biggest free-reach lever you have. Pick top 5 from #NBA/#NFL TL"),
        ("11:00 AM", "Reddit", "Value-add comment in r/sportsbook", "5 min", "Manual", "Never shill. Comment with insight; brand mentioned 1× max"),
        ("12:00 PM", "TikTok", "📊 Yesterday's result reveal (30-sec receipts focus)", "5 min", "Manual", "Receipts > picks. Trust = follower retention"),
        ("12:05 PM", "Instagram Reel", "Cross-post the 12pm TikTok", "30 sec", "Manual", ""),
        ("12:10 PM", "YouTube Short", "Cross-post", "30 sec", "Manual", ""),
        ("12:30 PM", "Twitter", "Yesterday's result thread (5 tweets)", "5 min", "Manual", "Lead with the miss if there was one. Trust play."),
        ("1:00 PM", "Twitter", "Quote-tweet contrarian Vegas line", "2 min", "Manual", "Polite disagreement. Don't beef."),
        ("2:00 PM", "FB Page", "Auto-cross-post from morning IG", "0", "Auto", "Use IG's Meta Business Suite cross-post toggle"),
        ("3:00 PM (Tue/Wed/Thu)", "LinkedIn (founder)", "Long-form take from Noa", "10 min", "Manual", "3×/week max. Algo punishes weekend posts."),
        ("4:00 PM", "Reddit", "1-2 more comments in r/nba / r/nfl / niche subs", "5 min", "Manual", ""),
        ("5:00 PM", "IG Story", "Pre-game live tracking begins", "1 min", "Manual", "BTS / 'Pick1 about to be tested live'"),
        ("6:00 PM", "TikTok", "🔮 Game-day prep OR tomorrow's sneak peek", "5 min", "Manual", "Save sneak peek hook for low-engagement days"),
        ("6:05 PM", "Instagram Reel", "Cross-post", "30 sec", "Manual", ""),
        ("7:00 PM", "Twitter", "Live game reactions during marquee match", "10 min", "Manual", "Real-time = algo priority. Don't miss."),
        ("9:00 PM", "TikTok", "🔥 Live reaction to key moment (if it happens)", "5 min", "Manual", "Trigger: any game with >$100M wager handle"),
        ("10:00 PM", "Twitter", "Night's results wrap", "3 min", "Manual", "Final tweet of the day = next day's setup"),
        ("11:00 PM", "IG Story", "Tomorrow's preview teaser", "1 min", "Manual", "Keep loyal followers coming back tomorrow"),
    ]

    for r, row in enumerate(schedule, start=6):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY_FONT
            cell.alignment = WRAP if c in (3, 6) else LEFT
        # Color-code Auto vs Manual
        auto_cell = ws.cell(row=r, column=5)
        if row[4] == "Auto":
            auto_cell.fill = PatternFill("solid", fgColor="C6EFCE")
        ws.row_dimensions[r].height = 28

    # ── Repurposing matrix ──
    rep_row = len(schedule) + 8

    ws.merge_cells(f"A{rep_row}:F{rep_row}")
    ws.cell(row=rep_row, column=1, value="REPURPOSING MATRIX — one piece, 6-8 outlets")
    ws.cell(row=rep_row, column=1).font = Font(name="Calibri", bold=True, size=12, color=BRAND_BG)
    ws.cell(row=rep_row, column=1).fill = PatternFill("solid", fgColor="EFEFEF")
    ws.row_dimensions[rep_row].height = 22

    rep_headers = ["Source content", "→ Outlet 1", "→ Outlet 2", "→ Outlet 3", "→ Outlet 4", "→ Outlet 5+"]
    for c, h in enumerate(rep_headers, start=1):
        ws.cell(row=rep_row + 1, column=c, value=h)
    style_header_row(ws, rep_row + 1, len(rep_headers))

    repurposing = [
        ("📹 1 TikTok (45 sec)", "IG Reel", "YouTube Short", "Twitter video", "FB Reel", "IG Story teaser + LinkedIn clip + quote-tweet"),
        ("📊 1 IG Carousel (10 slides)", "Twitter thread (1 slide/tweet)", "LinkedIn carousel", "FB carousel", "IG Story sequence (4-5 stories)", "Substack newsletter + YT community post"),
        ("📝 1 Twitter thread (20 tweets)", "Substack essay (consolidated)", "LinkedIn long-form", "IG carousel", "TikTok talking-head", "Email blast"),
        ("📷 1 Branded image (lineup card)", "Twitter image post", "IG feed post", "IG Story", "FB post", "LinkedIn image + Reddit /r/sportsbook image"),
        ("🎙️ 1 Twitter Space (60 min)", "Recap thread", "Clipped TikTok highlights", "Email recap", "Substack writeup", "Pinned on profile + boosted via X Premium"),
    ]
    for r, row in enumerate(repurposing, start=rep_row + 2):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY_FONT
            cell.alignment = WRAP
        ws.row_dimensions[r].height = 36

    # ── Hashtag library ──
    hash_row = rep_row + 2 + len(repurposing) + 2

    ws.merge_cells(f"A{hash_row}:F{hash_row}")
    ws.cell(row=hash_row, column=1, value="HASHTAG LIBRARY — copy-paste per platform")
    ws.cell(row=hash_row, column=1).font = Font(name="Calibri", bold=True, size=12, color=BRAND_BG)
    ws.cell(row=hash_row, column=1).fill = PatternFill("solid", fgColor="EFEFEF")
    ws.row_dimensions[hash_row].height = 22

    hashtags = [
        ("TikTok (3-5 per post)", "Wide reach", "#fyp #foryou #foryoupage #viral"),
        ("", "Sports general", "#sportstok #nbatok #nfltok #sportsbetting #parlay"),
        ("", "Sport-specific", "#nba #nfl #mlb #ufc #nhl #soccertok #worldcup #fifa"),
        ("", "Niche bettor", "#prizepicks #draftkings #fanduel #underdogfantasy"),
        ("", "AI angle", "#aitechnology #aitools #futuretech #aigeneration"),
        ("", "Custom brand", "#pick1 #beatpick1 #aisports #aisportsbetting"),
        ("Instagram (15-25 per post)", "Stack all above plus:", "#bettingtips #bettingexpert #picksdaily #sportspicks #freenba #freenfl #sportsanalytics #parlayoftheday"),
        ("Twitter/X (max 2 per post!)", "Use ONLY sport-specific:", "#NBA #NFL #NHL #WorldCup (etc.). Topical hashtags don't help on X anymore."),
        ("LinkedIn (3-5 per post)", "Professional angle", "#AI #MachineLearning #SportsTech #StartupLife #Founder"),
        ("Reddit (no hashtags, but flair)", "—", "Use sub-specific post flair. Tag posts properly or get removed."),
    ]
    hash_headers = ["Platform", "Category", "Hashtags (copy-paste)"]
    for c, h in enumerate(hash_headers, start=1):
        ws.cell(row=hash_row + 1, column=c, value=h)
    style_header_row(ws, hash_row + 1, len(hash_headers))

    for r, row in enumerate(hashtags, start=hash_row + 2):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY_FONT if c < 3 else Font(name="Consolas", size=10)
            cell.alignment = WRAP
        ws.row_dimensions[r].height = 30

    # ── Time-of-day map ──
    tod_row = hash_row + 2 + len(hashtags) + 2

    ws.merge_cells(f"A{tod_row}:F{tod_row}")
    ws.cell(row=tod_row, column=1, value="TIME-OF-DAY MAP (US sports audience, all times ET)")
    ws.cell(row=tod_row, column=1).font = Font(name="Calibri", bold=True, size=12, color=BRAND_BG)
    ws.cell(row=tod_row, column=1).fill = PatternFill("solid", fgColor="EFEFEF")
    ws.row_dimensions[tod_row].height = 22

    tod = [
        ("TikTok", "6-7am, 12-1pm, 6-8pm, 10-11pm", "Pre-coffee, lunch, post-work, pre-bed scrolls"),
        ("IG Reels", "11am, 6pm", "IG's daily algo evaluation windows"),
        ("IG Feed", "12pm, 7pm", "Lunch + dinner browse"),
        ("IG Stories", "Any (always-on)", "Algo doesn't down-rank story timing"),
        ("Twitter/X", "8-9am, 12pm, 5-6pm, 8-10pm", "Morning newsbite, lunch, commute, evening live-game"),
        ("Facebook", "1pm, 8pm", "FB algo skews older audience"),
        ("YouTube Shorts", "5-7pm, 9-10pm", "Evening watch session"),
        ("LinkedIn", "7-9am Tue/Wed/Thu only", "Algo punishes weekend founder posts"),
        ("Reddit", "8-11pm", "Peak browse hours for sports betting subs"),
    ]
    tod_headers = ["Platform", "Best post times", "Why"]
    for c, h in enumerate(tod_headers, start=1):
        ws.cell(row=tod_row + 1, column=c, value=h)
    style_header_row(ws, tod_row + 1, len(tod_headers))

    for r, row in enumerate(tod, start=tod_row + 2):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY_FONT
            cell.alignment = WRAP
        ws.row_dimensions[r].height = 22

    autosize(ws, [22, 26, 50, 12, 14, 50])


# ── Tab 11: Aggressive Tactics ───────────────────────────────────────────
# 50 free aggressive plays ranked by ROI. The CMO playbook — what to
# actually DO every day to maximize organic reach with $0 incremental.
def build_aggressive_tactics(wb):
    ws = wb.create_sheet("11 · Aggressive Tactics")
    ws.sheet_view.showGridLines = False

    # Section banner
    ws.merge_cells("A1:F1")
    ws["A1"] = "AGGRESSIVE TACTICS — 50 free moves ranked by ROI"
    ws["A1"].font = SECTION_FONT; ws["A1"].fill = SECTION_FILL; ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    ws["A2"] = ("Every move below: $0 spend. Listed by ROI (effort vs reach). "
                "Pick 10-15 that match your founder energy + audience. Execute daily. "
                "Tier 1 is highest leverage — do all of those. Tier 5 is guerrilla — pick 2-3 that fit.")
    ws["A2"].font = SUBTLE_FONT; ws["A2"].alignment = WRAP
    ws.row_dimensions[2].height = 36

    headers = ["#", "Tactic", "Tier", "Effort", "Expected reach", "How to execute"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=4, column=c, value=h)
    style_header_row(ws, 4, len(headers))
    ws.freeze_panes = "A5"

    tactics = [
        # ── TIER 1 — Highest leverage ──
        (1, "Reply-bomb sports Twitter", "1", "10 min/day", "5K-50K imps/day", "Every morning, reply to top 5 #NBA/#NFL tweets with Pick1's pre-game probability."),
        (2, "Trending sound jacking on TikTok", "1", "1 min/post", "10x algo boost", "Every TT uses a currently-trending sound. Check Tiktok's Creative Center daily."),
        (3, "Tag athletes when relevant", "1", "30 sec/post", "100K-1M imps if athlete engages", "When Pick1 nails a player prop, tag the player. 1 like from athlete = millions of impressions."),
        (4, "Tag sports media (Schefter, Woj, McAfee)", "1", "30 sec/post", "10K-1M imps", "When Pick1's prob matches a covered result, reply with screenshot tagging them."),
        (5, "Quote-tweet contrarian Vegas takes", "1", "2 min/day", "5K-50K imps", "Polite disagreement = reach. When a tout posts a hot take, QT with Pick1's actual model."),
        (6, "First-comment on big tweets", "1", "5 min/day", "5K-50K imps", "First reply under viral sports tweet often gets 10K+ impressions itself. Be fast."),
        (7, "Twitter Spaces (host weekly)", "1", "60 min/wk", "30-100 live + recording reach", "Friday 8pm: 'Weekly Picks Review'. $0. Recurring audience."),
        (8, "Streak content", "1", "1 min/day", "2K-10K imps/post", "Every Pick1 win extends the narrative. 'Pick1 is X-Y in the last Z days.' Update daily."),
        (9, "Counter-bets to public consensus", "1", "5 min/post", "5K-30K imps", "Explicitly bet against the public. Document. Win or lose, content."),
        (10, "Receipts-first content (own the misses)", "1", "5 min/loss", "Trust = follower retention", "Every wrong call gets its own owned-up post. 'We were wrong. Here's why.' Trust play."),
        # ── TIER 2 — Set-and-forget high-value ──
        (11, "Tweetdeck alerts on key sports words", "2", "30 min setup", "Be FIRST on breaking news", "Set up X Pro alerts on: Schefter, Woj, McAfee, ESPN, Bleacher Report."),
        (12, "Reddit value-add commenting", "2", "10 min/day", "Karma + brand affinity", "Be in r/sportsbook daily. Comment value, mention Pick1 1× max per 10 comments."),
        (13, "YouTube comment trading", "2", "10 min/day", "10K-50K imps/comment if pinned", "Top 10 sports YT channels — drop value-add comments daily."),
        (14, "DM 20 sports newsletter operators", "2", "2 hrs once", "1K-10K signups per inclusion", "'Free Pick1 tool' feature in their daily. $0 deal."),
        (15, "DM 10 small podcast hosts", "2", "2 hrs once", "1K-10K listeners each", "Sports betting podcasts (1K-10K listeners). Offer founder for 30-min interview."),
        (16, "TikTok duet-bait videos", "2", "5 min/post", "10x reach via duets", "'I bet $X on tonight's pick' → people duet with theirs. Each duet = free distribution."),
        (17, "TikTok 'wait for it' hooks", "2", "0 (built-in)", "2-3x algo boost", "Every TT starts with 'Wait for it...' or 'POV:'. Algo loves."),
        (18, "Comment-baiting captions", "2", "0 (built-in)", "Engagement signal", "'Drop a 🏀 if you're betting Knicks tonight'. Engagement = algo lift."),
        (19, "Engagement pod with 3-5 indie founders", "2", "5 min/day", "First-hour boost", "Informal group that likes each other's posts in first hour. Critical for X/IG algo."),
        (20, "Algorithm-friendly post lengths", "2", "0 (built-in)", "Higher completion %", "Twitter: 70-100 chars. TikTok: 30-60 sec. IG Reel: 15-30 sec."),
        # ── TIER 3 — Content formats that work ──
        (21, "Confidence-card images", "3", "10 min/day", "Shareable as wallpaper", "Branded squares with tonight's confidence %. Native to share."),
        (22, "Daily tracker screenshots", "3", "1 min/day", "Repeatable trust signal", "Every morning, post pick1.live/tracker showing the record."),
        (23, "AI vs Vegas graphics", "3", "5 min/post", "Visual + repeatable", "Side-by-side: Vegas line vs Pick1 probability."),
        (24, "Streak counters as graphics", "3", "2 min/day", "Daily content auto-generated", "Daily countdown of consecutive correct calls. Animatable too."),
        (25, "'How Pick1 thinks' carousels", "3", "30 min/post (weekly)", "Educational + trust-building", "Walk through one pick's logic. IG carousel format."),
        (26, "Pre-game vs post-game side-by-sides", "3", "5 min/post", "Receipts wall", "Pick1 said X, result was Y. Stack across the week."),
        (27, "Player-prop callouts", "3", "5 min/post", "Athlete may engage", "Tag specific players when relevant."),
        (28, "Sports-Twitter-bait takes", "3", "5 min/post", "Ratio'd by fans = free reach", "Sub-tweet popular bettor takes with model output. Designed to be ratio'd."),
        (29, "Live game reactions", "3", "10 min/game", "Real-time algo priority", "React in real-time. TikTok + Twitter especially."),
        (30, "End-of-night recap threads", "3", "5 min/night", "Trust + retention", "Every night, recap the day's calls in a Twitter thread."),
        # ── TIER 4 — Community engagement ──
        (31, "Reply to every DM within 1hr", "4", "Always-on", "Loyalty multiplier", "Loyalty multiplier. Sets retention pattern."),
        (32, "Pin best community comments", "4", "1 min/day", "Reciprocity", "Make commenters feel seen. They post more."),
        (33, "Weekly 'best community pick' feature", "4", "5 min/wk", "Reciprocity", "Spotlight a follower's pick on Pick1's account."),
        (34, "Birthday DMs to top engagers", "4", "5 min/day", "Cheap loyalty", "Pulled from analytics. Personal touch."),
        (35, "Friday AMA on X Spaces", "4", "30 min/wk", "30-100 live attendees", "Builds audience. $0."),
        (36, "Discord crashing (other servers)", "4", "30 min/wk", "Targeted audience", "Be in betting Discords, add value, mention Pick1 occasionally."),
        (37, "Telegram channel value-bombing", "4", "30 min/wk", "Niche but engaged", "Same in sports betting Telegram groups."),
        (38, "Sports forum posts (Covers, SBR, BetPop)", "4", "30 min/wk", "Old-school but loyal", "One value-add post/week per forum."),
        (39, "Reddit AMA in r/sportsbook (Day 16-18)", "4", "60 min once", "5K-50K signups", "Schedule one for late campaign. Pre-coordinate with mods."),
        (40, "Twitter Circle/Close Friends list", "4", "Always-on", "Loyalty hack", "Give 100 most engaged an exclusive list. Loyalty signal."),
        # ── TIER 5 — Guerrilla moves ──
        (41, "Live-tweet from sportsbook reply sections", "5", "10 min/day", "Targeted", "Be everywhere a sportsbook tweets. Reply with Pick1's model output."),
        (42, "Athlete birthday tweets", "5", "5 min/day", "Sometimes athlete replies", "Tag athletes on birthdays with their Pick1 stats. Sometimes go viral."),
        (43, "Sports anniversary tweets", "5", "5 min/wk", "Niche viral potential", "'X years ago today' with Pick1's retroactive probability."),
        (44, "Riding viral threads", "5", "Reactive", "Free if you hit it", "When sports thread goes viral, drop into replies with Pick1 data."),
        (45, "'What does the model say' reply template", "5", "30 min setup", "Builds anticipation", "Saved reply that fires when someone asks about a game."),
        (46, "Custom Pick1-win GIFs in reply", "5", "1 hr design", "Memorable replies", "Drop them in Twitter replies when relevant."),
        (47, "Submit custom emojis to communities", "5", "30 min total", "Brand affinity", "'Pick1 win' / 'Pick1 loss' emojis to relevant Discords/Slacks."),
        (48, "Pre-made meme stockpile", "5", "2 hrs once", "Cheap virality", "5-10 memes ready to fire when sports media has bad take."),
        (49, "Reaction reels to viral betting fails", "5", "5 min/reactive", "Trust contrast", "When a tout fails publicly, react with Pick1's actual call."),
        (50, "'This account uses Pick1' badge program", "5", "1 hr setup", "Tiny but additive", "Give 100 micro-creators a Pick1 verification badge for bio."),
    ]

    for r, row in enumerate(tactics, start=5):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY_FONT
            cell.alignment = WRAP if c == 6 else CENTER if c in (1, 3, 4) else LEFT
        # Tier color coding
        tier_cell = ws.cell(row=r, column=3)
        tier = str(row[2])
        if tier == "1":
            tier_cell.fill = PatternFill("solid", fgColor="C6EFCE")  # green - do all
        elif tier == "2":
            tier_cell.fill = PatternFill("solid", fgColor="DDEBF7")  # blue - high value
        elif tier == "3":
            tier_cell.fill = PatternFill("solid", fgColor="FFEB9C")  # yellow - formats
        elif tier == "4":
            tier_cell.fill = PatternFill("solid", fgColor="FFF3E0")  # orange - community
        elif tier == "5":
            tier_cell.fill = PatternFill("solid", fgColor="FCE4EC")  # pink - guerrilla
        ws.row_dimensions[r].height = 32

    autosize(ws, [5, 38, 7, 14, 26, 70])

    # ── CMO call-out at the bottom ──
    cmo_row = len(tactics) + 7

    ws.merge_cells(f"A{cmo_row}:F{cmo_row}")
    ws.cell(row=cmo_row, column=1, value="CMO FINAL CALL — if you can only pick 5")
    ws.cell(row=cmo_row, column=1).font = SECTION_FONT
    ws.cell(row=cmo_row, column=1).fill = SECTION_FILL
    ws.cell(row=cmo_row, column=1).alignment = CENTER
    ws.row_dimensions[cmo_row].height = 32

    cmo_picks = [
        ("1. Triple TikTok output (2 → 6 per day)",
         "Single biggest gap. Each extra TT = $0 cost, fresh FYP shot. Cap at 6 to avoid throttle."),
        ("2. Reply-bomb sports Twitter every morning",
         "10 min, free, biggest raw-reach lever you have. Sets daily algorithmic momentum."),
        ("3. Trade Lifetime Pro for 30+ micro-influencer shoutouts",
         "Not 20 — 30+. Volume strategy. Each in-kind trade = $0, ~$500 of bought equivalent."),
        ("4. Daily mini-sweeps stacked on Gleam (Mon-Sun)",
         "$75/week, 50K-500K additional impressions. Fresh hook for each day's content."),
        ("5. Weekly Twitter Space on Friday 8pm ET",
         "$0, recurring 30-100 live attendees who compound into loyal followers."),
    ]
    for r, (label, txt) in enumerate(cmo_picks, start=cmo_row + 1):
        ws.cell(row=r, column=1, value=label).font = Font(name="Calibri", bold=True, size=11, color=BRAND_BG)
        ws.merge_cells(f"B{r}:F{r}")
        ws.cell(row=r, column=2, value=txt).font = BODY_FONT
        ws.cell(row=r, column=2).alignment = WRAP
        ws.row_dimensions[r].height = 38


# ── Tab 12: $25K Active Plan ─────────────────────────────────────────────
# The CURRENT budget (authorized May 14, Day 2). Shows the actual spend
# plan, week-by-week cadence, expected outcomes, and immediate first
# moves. Tab 8 has the side-by-side scenarios; this tab is the live plan.
def build_active_plan(wb):
    ws = wb.create_sheet("12 · $25K Active Plan")
    ws.sheet_view.showGridLines = False

    # Section banner
    ws.merge_cells("A1:F1")
    ws["A1"] = "$25K ACTIVE BUDGET — current plan (authorized Day 2, May 14)"
    ws["A1"].font = SECTION_FONT; ws["A1"].fill = SECTION_FILL; ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 32

    # Intro
    ws.merge_cells("A2:F2")
    ws["A2"] = ("Total: $25,000 over 17 days (~$1,470/day). "
                "Expected: 8K-15K signups, 5K-15K IG followers, 8K-20K TikTok, 3-5 mid press, 2-4 viral pieces. "
                "Companion doc: launch/AT_25K.md")
    ws["A2"].font = SUBTLE_FONT; ws["A2"].alignment = WRAP
    ws.row_dimensions[2].height = 38

    # ── IMMEDIATE first moves banner ──
    ws.merge_cells("A4:F4")
    ws["A4"] = "🚨 IMMEDIATE FIRST MOVES — tonight (~10 min)"
    ws["A4"].font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    ws["A4"].fill = PatternFill("solid", fgColor="D32F2F")
    ws["A4"].alignment = CENTER
    ws.row_dimensions[4].height = 28

    first_moves = [
        ("1.", "Upgrade Gleam to Hobby plan — $39 — at https://gleam.io/billing"),
        ("2.", "Reply 'done' so Claude can autonomously upload hero image + add Viral Share + Visit-a-Page (TT/X) + Question entry"),
        ("3.", "Authorize Meta Ads payment method (Campaign 1 already set up at $50/day; scale to $350/day tomorrow morning)"),
    ]
    for r, (n, txt) in enumerate(first_moves, start=5):
        ws.cell(row=r, column=1, value=n).font = Font(name="Calibri", bold=True, size=12)
        ws.cell(row=r, column=1).alignment = CENTER
        ws.merge_cells(f"B{r}:F{r}")
        ws.cell(row=r, column=2, value=txt).font = BODY_FONT
        ws.cell(row=r, column=2).alignment = WRAP
        ws.row_dimensions[r].height = 26

    # ── Budget breakdown table ──
    bud_row = len(first_moves) + 6

    ws.merge_cells(f"A{bud_row}:F{bud_row}")
    ws.cell(row=bud_row, column=1, value="BUDGET BREAKDOWN — where every dollar goes")
    ws.cell(row=bud_row, column=1).font = Font(name="Calibri", bold=True, size=12, color=BRAND_BG)
    ws.cell(row=bud_row, column=1).fill = PatternFill("solid", fgColor="EFEFEF")
    ws.row_dimensions[bud_row].height = 22

    bud_headers = ["Category", "Line item", "$ amount", "Cadence", "Status", "Notes"]
    for c, h in enumerate(bud_headers, start=1):
        ws.cell(row=bud_row + 1, column=c, value=h)
    style_header_row(ws, bud_row + 1, len(bud_headers))

    budget = [
        # Paid social
        ("[PAID SOCIAL]", "", "", "", "", ""),
        ("Paid social", "Meta Ads — cold (Pixel-validated)", 6000, "$353/day × 17", "Pending auth", "At $1.50 CPL = 4K signups"),
        ("Paid social", "Meta Ads — retargeting", 2500, "$147/day × 17", "Pending auth", "3-7× lower CPL. Burns hot."),
        ("Paid social", "TikTok Spark Ads", 3000, "Reactive on >50K Reels", "Pending setup", "Only boost organic winners. $3K cap."),
        ("  → Subtotal", "Paid social", 11500, "", "", "~6-10K signups"),
        # Influencer
        ("[INFLUENCER]", "", "", "", "", ""),
        ("Influencer", "1 × mid-tier sports creator (300-500K)", 3000, "One-off", "Pending outreach Day 4", "Single biggest single-shot reach. Vet engagement >2%."),
        ("Influencer", "5-7 × micro creators (50-150K)", 2500, "Day 7-12 rollout", "Pending outreach Day 7", "$300-500 each. Mix paid + product trade."),
        ("Influencer", "Lifetime Pro trades (15-20 micros)", 0, "Always-on", "Pending Day 7", "From INFLUENCER_OUTREACH.md. In-kind only."),
        ("  → Subtotal", "Influencer", 5500, "", "", "200K-1M impressions"),
        # Production
        ("[CONTENT PRODUCTION]", "", "", "", "", ""),
        ("Production", "Upwork content editor (18 days)", 500, "$28/day × 18", "Pending Day 5 hire", "Cross-posting + editing while founder records"),
        ("Production", "1 × day production shoot (street TikTok)", 1500, "Day 10 (May 22)", "Pending", "NYC or LA. Creator + lighting + edit. 4-6 pieces."),
        ("Production", "Canva Pro Team + Pond5 + motion design", 500, "One-off", "Pending Day 6", "Reusable asset library"),
        ("  → Subtotal", "Production", 2500, "", "", "Quality + volume both up"),
        # Sweeps & viral
        ("[SWEEPS & VIRAL]", "", "", "", "", ""),
        ("Sweeps", "🎯 Gleam Hobby plan ($39 × 1 mo)", 40, "Tonight", "★ FIRST ACTION", "Unlocks Feature Media + Viral Share + Repeat Action Limit"),
        ("Sweeps", "Sweepstakes prize stack upgrade", 2000, "Day 12 announce", "Pending", "Add $1.5K cash 2nd prize. Mid-campaign news hook."),
        ("Sweeps", "SparkLoop newsletter cross-promo", 1500, "Day 8 submit", "Pending", "750-1500 signups expected."),
        ("Sweeps", "Daily Drop on Twitter (7 × $40-50)", 300, "$43/day × 7", "Pending Day 16", "Week-of-launch micro-prizes."),
        ("Sweeps", "KingSumo Pro (free tier sufficient)", 0, "Day 10 launch", "Pending", "Free tier supports 1,000 entries — enough for our scale."),
        ("  → Subtotal", "Sweeps & viral", 3840, "", "", "More entries + better mechanics"),
        # PR
        ("[PR & LAUNCH]", "", "", "", "", ""),
        ("PR", "ProductHunt paid hunter (Day 18 launch)", 500, "One-off", "Pending Day 17 outreach", "Experienced hunter + animated demo + day-of comms."),
        ("PR", "Press kit (photography + bio + media training)", 500, "Day 14 session", "Pending booking", "Professional shots for press intake."),
        ("  → Subtotal", "PR", 1000, "", "", "Founder story drop amplification"),
        # Tooling
        ("[TOOLING & INFRA]", "", "", "", "", ""),
        ("Tooling", "Resend Pro + Vercel Pro + Supabase Pro", 200, "Day 5 upgrade", "Pending auth", "Needed once waitlist crosses 5K"),
        ("  → Subtotal", "Tooling", 200, "", "", ""),
        # Contingency
        ("[CONTINGENCY]", "", "", "", "", ""),
        ("Contingency", "Day 10 reallocation reserve", 460, "Day 10 decision", "Held", "★ Critical lever. Whatever's winning gets this."),
        # ── TOTAL ──
        ("", "", "", "", "", ""),
        ("TOTAL", "", 25000, "", "", "$1,470/day equivalent"),
    ]

    for r_idx, row in enumerate(budget, start=bud_row + 2):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c, value=v)
            cell.font = BODY_FONT
            cell.alignment = WRAP if c == 6 else (CENTER if c in (3, 4) else LEFT)
            if c == 3 and isinstance(v, (int, float)):
                cell.number_format = "$#,##0"

            # Section header rows
            if str(row[0]).startswith("[") and str(row[1]) == "":
                cell.font = Font(name="Calibri", bold=True, size=11, color=BRAND_BG)
                cell.fill = PatternFill("solid", fgColor="F4F4F4")
            # Subtotal rows
            elif str(row[0]).strip().startswith("→"):
                cell.font = Font(name="Calibri", bold=True, size=10, color="333333")
                cell.fill = PatternFill("solid", fgColor="EFEFEF")
            # TOTAL row
            elif row[0] == "TOTAL":
                cell.font = Font(name="Calibri", bold=True, size=12, color=BRAND_BG)
                cell.fill = HEADER_FILL
            # First action highlight
            if "FIRST ACTION" in str(v):
                cell.fill = PatternFill("solid", fgColor="FFE082")
                cell.font = Font(name="Calibri", bold=True, size=10)

        ws.row_dimensions[r_idx].height = 28

    autosize(ws, [16, 38, 11, 22, 18, 50])

    # ── Week-by-week cadence ──
    week_row = bud_row + 2 + len(budget) + 3

    ws.merge_cells(f"A{week_row}:F{week_row}")
    ws.cell(row=week_row, column=1, value="WEEK-BY-WEEK PAID CADENCE")
    ws.cell(row=week_row, column=1).font = SECTION_FONT
    ws.cell(row=week_row, column=1).fill = SECTION_FILL
    ws.cell(row=week_row, column=1).alignment = CENTER
    ws.row_dimensions[week_row].height = 32

    week_headers = ["Day", "Date", "Paid action", "$ spent (cumulative)", "Notes"]
    for c, h in enumerate(week_headers, start=1):
        ws.cell(row=week_row + 1, column=c, value=h)
    style_header_row(ws, week_row + 1, len(week_headers))

    cadence = [
        # Week 1: Ramp up
        ("[Week 1: RAMP UP]", "", "", "", ""),
        (2, "May 14", "Gleam Hobby upgrade ($39) — TONIGHT", 39, "★ First action. Unlocks everything else."),
        (3, "May 15", "Meta Ads cold scaled to $350/day starts", 389, "Existing Campaign 1 just gets the bigger budget"),
        (3, "May 15", "TikTok Spark Ads $175/day reactive starts", 389, "Only fires on Reels >50K views"),
        (4, "May 16", "DM 5 mid-tier creators with $3K paid offer", 389, "No spend yet — outreach phase"),
        (5, "May 17", "Upwork content editor hired ($500 one-off)", 889, "18-day gig"),
        (5, "May 17", "Infra upgrades ($200): Resend/Vercel/Supabase Pro", 1089, ""),
        (6, "May 18", "Canva Pro Team + Pond5 ($500 one-off)", 1589, ""),
        (7, "May 19", "DM 15-20 micros with $300-500 paid offer", 1589, "Outreach phase"),
        (8, "May 20", "SparkLoop submission ($1,500 cap)", 3089, "24-72hr approval"),
        ("  → Week 1 spend", "", "", 3089, "$181/day avg — mostly setup"),
        # Week 2: Amplify
        ("[Week 2: AMPLIFY]", "", "", "", ""),
        (9, "May 21", "Mid-tier creator post drops ($3K)", 6089, "Single biggest line item lands"),
        (9, "May 21", "First micro creator posts drop ($1,500)", 7589, "3-4 micros over Day 9-12"),
        (10, "May 22", "Production shoot day ($1,500)", 9089, "Street TikTok content"),
        (10, "May 22", "★ HARD REALLOCATION ($460 contingency)", 9549, "★★★ Critical decision moment"),
        (11, "May 23", "More micro posts drop ($1,000)", 10549, "Remaining 2-3 micros"),
        (12, "May 24", "Sweepstakes prize bump announcement ($2,000)", 12549, "Cash 2nd prize added"),
        (14, "May 26", "Founder story drop + paid amplification ($500)", 13049, "Boost the best story-drop video"),
        (14, "May 26", "Press kit photography session ($500)", 13549, "For founder story wave press intake"),
        ("  → Week 2 spend", "", "", 13549, "$1,494/day avg — heaviest week"),
        # Week 3: Final push
        ("[Week 3: FINAL PUSH]", "", "", "", ""),
        (16, "May 28", "Daily Drop on Twitter starts ($43/day × 7 = $300)", 13849, "$300 across week"),
        (17, "May 29", "Meta retargeting full-throttle ($147/day)", 16349, "Cumulative retargeting line keeps adding"),
        (18, "May 30", "ProductHunt paid hunter engaged ($500)", 16849, "Day-of comms support"),
        (18, "May 30", "TikTok Spark Ads final push ($1,000 of $3K cap)", 17849, "Use remaining boost budget on top performers"),
        (19, "May 31", "Final-day Meta ad surge ($500 extra)", 18349, "Last-day FOMO"),
        ("  → Week 3 spend", "", "", 18349, "—"),
        # Cumulative running through May 31
        ("", "", "", "", ""),
        ("Always-on through May 31", "Meta cold ($350/day × 17)", "", 25000, "Cumulative total all in"),
    ]

    for r_idx, row in enumerate(cadence, start=week_row + 2):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c, value=v)
            cell.font = BODY_FONT
            cell.alignment = WRAP if c == 5 else CENTER
            if c == 4 and isinstance(v, (int, float)):
                cell.number_format = "$#,##0"

            # Section header rows
            if str(row[0]).startswith("["):
                cell.font = Font(name="Calibri", bold=True, size=11, color=BRAND_BG)
                cell.fill = PatternFill("solid", fgColor="F4F4F4")
            elif str(row[0]).strip().startswith("→"):
                cell.font = Font(name="Calibri", bold=True, size=10, color="333333")
                cell.fill = PatternFill("solid", fgColor="EFEFEF")

        ws.row_dimensions[r_idx].height = 24

    # ── Expected outcomes ──
    out_row = week_row + 2 + len(cadence) + 3

    ws.merge_cells(f"A{out_row}:F{out_row}")
    ws.cell(row=out_row, column=1, value="EXPECTED OUTCOMES — $1K plan vs $25K active vs $100K hypothetical")
    ws.cell(row=out_row, column=1).font = SECTION_FONT
    ws.cell(row=out_row, column=1).fill = SECTION_FILL
    ws.cell(row=out_row, column=1).alignment = CENTER
    ws.row_dimensions[out_row].height = 32

    out_headers = ["Metric", "$1K plan", "$25K plan (CURRENT)", "$100K plan", "Lift ($25K vs $1K)", "Notes"]
    for c, h in enumerate(out_headers, start=1):
        ws.cell(row=out_row + 1, column=c, value=h)
    style_header_row(ws, out_row + 1, len(out_headers))

    outcomes = [
        ("Waitlist signups", "1.5-2.5K", "8K-15K", "25K-50K", "5-6×", "Linear scaling on paid channels"),
        ("Cost per signup (blended)", "$0.40-0.67", "$1.67-3.13", "$2-4", "↑ (pay for scale)", "Higher CPL is normal at scale"),
        ("IG followers gained", "0.5-1.5K", "5K-15K", "20K-50K", "10×", "Paid TikTok cross-pollinates"),
        ("TikTok followers gained", "1K-3K", "8K-20K", "30K-75K", "6.7×", "Spark Ads + production shoot"),
        ("Twitter followers gained", "200-500", "2K-5K", "8K-20K", "10×", "Founder story drop + Show HN"),
        ("Press placements", "1 small", "3-5 mid", "5-10 + tier-1", "3-5×", "Press kit + paid hunter"),
        ("Viral pieces (>100K views)", "0-1", "2-4", "4-8", "2-4×", "Production shoot designed for this"),
        ("Brand recall (post-launch survey)", "<5%", "10-20%", "25-40%", "2-4×", "Compounding effect"),
        ("Lifetime Pro conversions", "15-30", "80-300", "250-1.5K", "5-10×", "$30 launch promo, $150 regular"),
    ]
    for r_idx, row in enumerate(outcomes, start=out_row + 2):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c, value=v)
            cell.font = BODY_FONT
            cell.alignment = WRAP if c == 6 else CENTER
            # Highlight $25K column
            if c == 3:
                cell.font = Font(name="Calibri", bold=True, size=10, color=BRAND_BG)
                cell.fill = PatternFill("solid", fgColor="E8F5E9")
        ws.row_dimensions[r_idx].height = 26

    # ── Risk callouts at bottom ──
    risk_row = out_row + 2 + len(outcomes) + 3

    ws.merge_cells(f"A{risk_row}:F{risk_row}")
    ws.cell(row=risk_row, column=1, value="RISK CALLOUTS")
    ws.cell(row=risk_row, column=1).font = SECTION_FONT
    ws.cell(row=risk_row, column=1).fill = SECTION_FILL
    ws.cell(row=risk_row, column=1).alignment = CENTER
    ws.row_dimensions[risk_row].height = 32

    risks = [
        ("Meta CPL > $3 risk",
         "If Day 7 CPL on cold > $3, kill cold campaign immediately. Redeploy budget to retargeting + influencer. Don't slow-drift."),
        ("Mid-tier creator dud risk",
         "$3K mid-tier post is binary — if audience doesn't convert, that's 12% of budget gone. Vet engagement rate (>2% required) before signing. Read 5 most recent posts to confirm they're not running competing betting promos."),
        ("Production shoot weather/logistics",
         "$1.5K street TikTok shoot can rain out. Block 2 backup dates. Or shift to studio if weather is bad."),
        ("Gleam Hobby vs Pro at scale",
         "If Viral Share entries explode past 1K participants, consider one-time $99 Pro upgrade. Decision point: Day 10. Total budget impact: $99 from contingency."),
        ("Algorithm dependence",
         "At $25K we're more exposed to platform algo changes mid-campaign. Diversification across Meta + TikTok + organic is the hedge."),
    ]
    for r, (label, txt) in enumerate(risks, start=risk_row + 1):
        ws.cell(row=r, column=1, value=label).font = Font(name="Calibri", bold=True, size=11, color=BRAND_BG)
        ws.merge_cells(f"B{r}:F{r}")
        ws.cell(row=r, column=2, value=txt).font = BODY_FONT
        ws.cell(row=r, column=2).alignment = WRAP
        ws.row_dimensions[r].height = 38


# ── README tab (first one) ───────────────────────────────────────────────
def build_readme(wb):
    ws = wb.create_sheet("0 · README", 0)  # insert first
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws["A1"] = "PICK1 · MASTER TRACKER"
    ws["A1"].font = Font(name="Calibri", bold=True, size=22, color=WHITE)
    ws["A1"].fill = SECTION_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 44

    lines = [
        ("How to use this", "header"),
        ("This sheet is the EXECUTION log. The strategy doc (launch/MASTER_STRATEGY.md) is the plan; this is the daily-update tracker.", "body"),
        ("", "body"),
        ("Update cadence: every night before bed. Takes ~3 min.", "body"),
        ("Sundays: full audit + reallocation decisions for the week ahead.", "body"),
        ("", "body"),
        ("⭐ Start here every day", "header"),
        ("Tab 9 — Daily Strategy — has the day-by-day execution plan for May 14 → May 31. Auto-cron vs. you-post vs. outreach vs. activations, with daily + cumulative signup targets. Open this tab every morning.", "body"),
        ("", "body"),
        ("The 9 tabs", "header"),
        ("1. KPI Dashboard         — top-level daily numbers (signups, CPL, etc.) [your SCOREBOARD]", "body"),
        ("2. Content Calendar      — every post planned + posted, with perf", "body"),
        ("3. Ad Performance        — Meta + TikTok Spark per-day per-ad", "body"),
        ("4. Channel ROI           — which channel is winning, % of total (42 channels tracked)", "body"),
        ("5. Activations           — the 22 viral activations + state (web3 quests, Founder Pass, etc.)", "body"),
        ("6. Influencer Outreach   — the 20-DM micro-trade program", "body"),
        ("7. Pick Performance      — daily pick record (the receipts wall)", "body"),
        ("8. $100K Scenario        — what marketing spend would look like at $1K vs $25K vs $100K", "body"),
        ("9. Daily Strategy ⭐    — day-by-day execution plan May 14 → May 31 (THE plan)", "body"),
        ("10. Content Engine      — daily content blueprint + repurposing matrix + hashtag library + times", "body"),
        ("11. Aggressive Tactics  — 50 free moves ranked by ROI (CMO playbook, tier-coded)", "body"),
        ("12. $25K Active Plan ⭐ — CURRENT BUDGET. Full allocation + cadence + outcomes. Open this with Tab 9.", "body"),
        ("", "body"),
        ("North Star", "header"),
        ("1,500 waitlist signups by May 31. Stretch target: 2,500.", "body"),
        ("Budget: $1,000 total. CPL target: < $1.50.", "body"),
        ("", "body"),
        ("Quick references", "header"),
        ("Strategy doc:   launch/MASTER_STRATEGY.md", "body"),
        ("Site:           https://pick1.live", "body"),
        ("FB Page:        https://www.facebook.com/profile.php?id=61589631453496", "body"),
        ("Resend domain:  pick1.live (verified)", "body"),
        ("Supabase proj:  lgnjawngkiamlngcffrk (eu-central-2)", "body"),
    ]
    for i, (text, style) in enumerate(lines, start=3):
        ws.cell(row=i, column=1, value=text)
        if style == "header":
            ws.cell(row=i, column=1).font = Font(name="Calibri", bold=True, size=14, color=BRAND_BG)
            ws.row_dimensions[i].height = 22
        else:
            ws.cell(row=i, column=1).font = Font(name="Calibri", size=11)
    ws.column_dimensions["A"].width = 110


def main():
    wb = Workbook()
    # Remove the default sheet so we control insertion order
    default = wb.active
    wb.remove(default)

    build_kpi_dashboard(wb)
    build_content_calendar(wb)
    build_ad_performance(wb)
    build_channel_roi(wb)
    build_activations(wb)
    build_influencer_outreach(wb)
    build_pick_performance(wb)
    build_scenario_budget(wb)
    build_daily_strategy(wb)
    build_content_engine(wb)
    build_aggressive_tactics(wb)
    build_active_plan(wb)
    build_readme(wb)  # inserted at position 0

    out = Path(__file__).parent / "MASTER_TRACKER.xlsx"
    wb.save(out)
    print(f"✓ Wrote {out}  ({len(wb.sheetnames)} tabs)")


if __name__ == "__main__":
    main()
